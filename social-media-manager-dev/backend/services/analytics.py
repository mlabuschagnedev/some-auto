from __future__ import annotations

import os
import re
import difflib
import unicodedata
from io import BytesIO
from pathlib import Path
from threading import Lock, Thread
from typing import Callable

from flask import send_file

from ..models import (
    AccountInsightSnapshot,
    InstagramFollowerSnapshot,
    Page,
    PlatformPostReference,
    Post,
    PostInsightSnapshot,
    SocialAccount,
    SocialInsight,
    db,
)
from ..routes.common import Any, jsonify, request, utcnow
from ..media import build_local_media_url
from .. import app as core

API_TIMEOUT_SECONDS = core.API_TIMEOUT_SECONDS
SOCIAL_INSIGHTS_META_API_VERSION = core.SOCIAL_INSIGHTS_META_API_VERSION
SOCIAL_INSIGHTS_META_API_FALLBACK_VERSION = core.SOCIAL_INSIGHTS_META_API_FALLBACK_VERSION
SOCIAL_INSIGHTS_REFRESH_INTERVAL_SECONDS = core.SOCIAL_INSIGHTS_REFRESH_INTERVAL_SECONDS
SOCIAL_INSIGHTS_MIN_REFRESH_SECONDS = core.SOCIAL_INSIGHTS_MIN_REFRESH_SECONDS
SOCIAL_INSIGHTS_ACCOUNT_PACE_SECONDS = core.SOCIAL_INSIGHTS_ACCOUNT_PACE_SECONDS
datetime = core.datetime
timedelta = core.timedelta
json = core.json
logger = core.logger
requests = core.requests
time = core.time
uuid = core.uuid
func = core.func
or_ = core.or_

REMOVED_INSIGHT_METRICS = {
    "fan_count",
    "fans",
    "online_followers",
    "page_fans",
    "page_fans_city",
    "page_fans_country",
    "page_fans_locale",
    "page_fan_adds",
    "page_fan_adds_unique",
    "page_fan_adds_by_paid_non_paid_unique",
    "page_fan_removes",
    "page_fan_removes_unique",
    "page_follows",
}

FACEBOOK_DAILY_METRIC_GROUPS = {
    "views": [
        {"metric": "page_media_view", "params": {"period": "day"}},
    ],
    "reach": [
        {"metric": "page_impressions_unique", "params": {"period": "day"}},
    ],
    "visits": [
        {"metric": "page_views_total", "params": {"period": "day"}},
    ],
    "followers": [
        {"metric": "page_follows", "params": {"period": "day"}},
    ],
    "engagement": [
        {"metric": "page_post_engagements", "params": {"period": "day"}},
    ],
    "reactions": [
        {"metric": "page_actions_post_reactions_total", "params": {"period": "day"}},
    ],
}

IG_USER_FIELD_METRICS = {"followers_count", "media_count"}
IG_STANDARD_SERIES_METRICS: set[str] = set()
IG_TOTAL_VALUE_METRICS = {"views", "reach", "profile_views", "accounts_engaged", "total_interactions"}
TOTAL_VALUE_INSIGHT_METRICS = IG_TOTAL_VALUE_METRICS
DISCOVERED_TOTAL_VALUE_INSIGHT_METRICS: set[str] = set()
GRAPH_VERSION_FALLBACK_CACHE_SECONDS = 600
GRAPH_VERSION_UNAVAILABLE_UNTIL: dict[str, float] = {}

INSTAGRAM_DAILY_METRIC_GROUPS = {
    "views": [
        {"metric": "views", "params": {"period": "day"}},
    ],
    "reach": [
        {"metric": "reach", "params": {"period": "day"}},
    ],
    "visits": [
        {"metric": "profile_views", "params": {"period": "day"}},
    ],
    "engagement": [
        {"metric": "accounts_engaged", "params": {"period": "day"}},
        {"metric": "total_interactions", "params": {"period": "day"}},
    ],
}

ACCOUNT_OBJECT_FIELDS = {
    "facebook": ["name", "fan_count", "followers_count"],
    "instagram": ["name", "username", "followers_count", "media_count"],
}

ACCOUNT_OBJECT_FIELD_METRICS = {
    "facebook": {
        "followers_count": "followers",
        "fan_count": "followers",
    },
    "instagram": {
        "followers_count": "followers",
        "media_count": "media_count",
    },
}

OPTIONAL_DAILY_METRICS = {
    "facebook": {"views", "reach", "visits", "followers", "engagement", "reactions"},
    "instagram": {"followers", "engagement", "visits", "reach", "views"},
}

ZERO_WHEN_EMPTY_DAILY_METRICS = {
    "facebook": set(),
}

REPORT_TEMPLATE_PATH_ENV = "MARKETING_REPORT_TEMPLATE_PATH"
DEFAULT_REPORT_TEMPLATE_PATH = r"E:\Marcel\Downloads\MSS Marketing Report 2026.xlsx"
GOOGLE_REPORT_CREDENTIALS_PATH_ENV = "GOOGLE_REPORT_CREDENTIALS_PATH"
GOOGLE_APPLICATION_CREDENTIALS_ENV = "GOOGLE_APPLICATION_CREDENTIALS"
GOOGLE_REPORT_SPREADSHEET_ID_ENV = "GOOGLE_REPORT_SPREADSHEET_ID"
GOOGLE_CAMPAIGN_SPREADSHEET_ID_ENV = "GOOGLE_CAMPAIGN_SPREADSHEET_ID"
DEFAULT_GOOGLE_REPORT_CREDENTIALS_PATH = r"E:\Marcel\Downloads\some-auto-480808-ae53fc67caf5.json"
DEFAULT_GOOGLE_REPORT_SPREADSHEET_ID = "1wYQ1ezEkm1IauemeggSnhzofkeyS3ER59sqw1ml3qsE"
DEFAULT_GOOGLE_CAMPAIGN_SPREADSHEET_ID = "1QQM1gBKBZxG3Y4A2A_EUlXRaExlfy_GQc24kxbAO_Js"
GOOGLE_SHEETS_SCOPES = ("https://www.googleapis.com/auth/spreadsheets",)
REPORT_IGNORED_SHEET_KEYS = {"onderengele"}
REPORT_SHEET_PAGE_ALIASES = {
    "udtrucksklerksdorp": "UD Trucks KLD",
    "udtrucksupington": "UD Trucks UPT",
}
REPORT_CAMPAIGN_SHEET_ALIASES = {
    "udtrucksklerksdorp": "UD Trucks KLD",
    "udtrucksupington": "UD Trucks UPT",
}
REPORT_POST_CONTENT_CLEAR_ROWS = 80
CAMPAIGN_POST_MATCH_MIN_SCORE = 0.85
CAMPAIGN_POST_MATCH_PAD_DAYS = 10
REPORT_MONTH_KEYS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}
REPORT_PLATFORM_KEYS = {
    "facebook": "facebook",
    "instagram": "instagram",
}
REPORT_SECTION_BOUNDARY_KEYS = {
    "facebook",
    "instagram",
    "linkedin",
    "google",
    "googleads",
    "facebookads",
    "instagramads",
    "linkedinads",
    "postcontent",
    "website",
    "seo",
}
REPORT_METRIC_KEYS = {
    "views": "views",
    "reach": "reach",
    "viewers": "reach",
    "reachviewers": "reach",
    "interactions": "engagement",
    "contentinteractions": "engagement",
    "engagement": "engagement",
    "visits": "visits",
    "followers": "followers",
    "totalfollowers": "followers",
    "lifetimefollowers": "followers",
    "totallifetimefollowers": "followers",
}
REPORT_SNAPSHOT_METRICS = {"followers", "media_count"}
REPORT_NON_ADDITIVE_ACTIVITY_METRICS = {"reach"}

REPORT_FACEBOOK_METRIC_GROUPS = {
    "views": [
        {"metric": "page_media_view", "params": {"period": "day", "metric_type": "total_value"}},
    ],
    "reach": [
        {"metric": "page_total_media_view_unique", "params": {"period": "day", "metric_type": "total_value"}},
        {"metric": "page_impressions_unique", "params": {"period": "days_28"}},
    ],
    "engagement": [
        {"metric": "page_post_engagements", "params": {"period": "day", "metric_type": "total_value"}},
    ],
    "visits": [
        {"metric": "page_views_total", "params": {"period": "day", "metric_type": "total_value"}},
    ],
    "followers": [
        {"metric": "page_follows", "params": {"period": "day"}},
    ],
}

REPORT_INSTAGRAM_METRIC_GROUPS = {
    "views": [
        {"metric": "views", "params": {"period": "day", "metric_type": "total_value"}, "max_days": 30},
    ],
    "reach": [
        {"metric": "reach", "params": {"period": "day", "metric_type": "total_value"}, "max_days": 30},
    ],
    "visits": [
        {"metric": "profile_views", "params": {"period": "day", "metric_type": "total_value"}, "max_days": 30},
    ],
    "engagement": [
        {"metric": "accounts_engaged", "params": {"period": "day", "metric_type": "total_value"}, "max_days": 30},
        {"metric": "total_interactions", "params": {"period": "day", "metric_type": "total_value"}, "max_days": 30},
    ],
}

POST_INSIGHT_METRICS = {
    "facebook": [
        "views",
        "reach",
        "engagement",
        "comments",
        "shares",
        "reactions",
        "clicks",
    ],
    "instagram": [
        "reach",
        "views",
        "likes",
        "comments",
        "shares",
        "saved",
        "total_interactions",
    ],
}
POST_INSIGHTS_LOOKBACK_DAYS = 60
INSTAGRAM_MEDIA_FIELDS = "id,caption,media_type,media_url,thumbnail_url,permalink,timestamp,like_count,comments_count"
FACEBOOK_POST_FIELDS = (
    "id,message,story,created_time,permalink_url,full_picture,shares,"
    "comments.limit(0).summary(true),reactions.limit(0).summary(true)"
)
FACEBOOK_POST_INSIGHT_METRIC_PRIORITY = {
    "views": [
        "post_media_view",
        "post_impressions",
    ],
    "reach": [
        "post_total_media_view_unique",
        "post_impressions_unique",
    ],
    "clicks": ["post_clicks"],
    "reactions": ["post_reactions_by_type_total"],
}
FACEBOOK_POST_INSIGHT_PARAM_OPTIONS = ({"period": "lifetime"}, {})

InsightMetricCandidate = dict[str, Any]
InsightRunContext = dict[str, Any]
AnalyticsProgressCallback = Callable[[int, int, SocialAccount], None]

_analytics_refresh_lock = Lock()
_analytics_refresh_state: dict[str, Any] = {
    "id": None,
    "status": "idle",
    "message": "No analytics refresh is running.",
    "account_id": None,
    "started_at": None,
    "finished_at": None,
    "progress_current": 0,
    "progress_total": 0,
    "result": None,
    "error": None,
}


def _safe_error(error: Exception | str) -> str:
    text = str(error)
    if "access_token" in text:
        return "Meta API request failed. Check account permissions or token health."
    try:
        payload = json.loads(text)
    except Exception:
        payload = None
    if isinstance(payload, dict):
        message = str(payload.get("message") or "")
        code = payload.get("code")
        subcode = payload.get("error_subcode") or payload.get("subcode")
        if code == 10 or "does not have permission" in message.lower():
            suffix = f" (code {code}{f', subcode {subcode}' if subcode else ''})" if code else ""
            return f"Meta permission unavailable for this insight metric. Check app permissions and account access.{suffix}"
        if message:
            suffix = f" (code {code}{f', subcode {subcode}' if subcode else ''})" if code else ""
            return f"{message}{suffix}"[:500]
    return text[:500]


def _date_range(days: int = 30) -> tuple[datetime, datetime]:
    end = utcnow()
    start = end - timedelta(days=max(days, 1))
    return start, end


def _start_of_day(value: datetime) -> datetime:
    return value.replace(hour=0, minute=0, second=0, microsecond=0)


def _end_of_day(value: datetime) -> datetime:
    return value.replace(hour=23, minute=59, second=59, microsecond=0)


def _graph_until_date(end: datetime) -> str:
    return (_start_of_day(end) + timedelta(days=1)).strftime("%Y-%m-%d")


def _insight_date_windows(start: datetime, end: datetime, *, max_days: int = 30) -> list[tuple[datetime, datetime]]:
    if end < start:
        return []
    window_days = max(int(max_days or 30), 1)
    final_day = _start_of_day(end)
    cursor = _start_of_day(start)
    windows: list[tuple[datetime, datetime]] = []
    while cursor <= final_day:
        chunk_end_day = min(cursor + timedelta(days=window_days - 1), final_day)
        windows.append((cursor, _end_of_day(chunk_end_day)))
        cursor = chunk_end_day + timedelta(days=1)
    return windows


def _parse_refresh_range_args() -> tuple[datetime, datetime]:
    default_start, default_end = _date_range(POST_INSIGHTS_LOOKBACK_DAYS)
    range_key = (request.args.get("range") or "").strip().lower()
    start: datetime | None = None
    end: datetime | None = None
    if range_key == "custom":
        try:
            start = core.parse_iso_datetime(f"{request.args.get('start')}T00:00:00") if request.args.get("start") else None
        except Exception:
            start = None
        try:
            end = core.parse_iso_datetime(f"{request.args.get('end')}T23:59:59") if request.args.get("end") else None
        except Exception:
            end = None
    elif range_key in {"7d", "30d", "60d"}:
        days = int(range_key[:-1])
        end = utcnow()
        start = end - timedelta(days=days)
    elif range_key == "month":
        end = utcnow()
        start = end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    start = start or default_start
    end = end or default_end
    if end < start:
        raise RuntimeError("Refresh end date must be after the start date.")
    return start, end


def _metric_names_for_platform(platform: str) -> list[str]:
    if platform == "facebook":
        return list(FACEBOOK_DAILY_METRIC_GROUPS.keys())
    if platform == "instagram":
        return list(INSTAGRAM_DAILY_METRIC_GROUPS.keys())
    return []


def _daily_metric_groups_for_platform(platform: str) -> dict[str, list[InsightMetricCandidate]]:
    if platform == "facebook":
        return FACEBOOK_DAILY_METRIC_GROUPS
    if platform == "instagram":
        return INSTAGRAM_DAILY_METRIC_GROUPS
    return {}


def _account_insight_candidate_params(platform: str, graph_metric: str, candidate: InsightMetricCandidate) -> dict[str, Any]:
    params = dict(candidate.get("params") or {})
    if platform == "instagram" and graph_metric in (TOTAL_VALUE_INSIGHT_METRICS | DISCOVERED_TOTAL_VALUE_INSIGHT_METRICS):
        params["metric_type"] = "total_value"
    return params


def _field_metric_groups_for_platform(platform: str) -> dict[str, list[str]]:
    return {}


def _friendly_metric_label(metric_name: str | None) -> str:
    labels = {
        "views": "Views",
        "page_media_view": "Views",
        "page_views_total": "Visits",
        "page_impressions_unique": "Reach",
        "page_total_media_view_unique": "Viewers",
        "engagement": "Content interactions",
        "page_post_engagements": "Content interactions",
        "followers": "Followers",
        "reach": "Reach",
        "visits": "Visits",
        "media_count": "Media count",
        "profile_views": "Visits",
        "accounts_engaged": "Content interactions",
        "reactions": "Reactions",
        "likes": "Likes",
        "comments": "Comments",
        "shares": "Shares",
        "saved": "Saves",
        "total_interactions": "Total interactions",
    }
    normalized = str(metric_name or "").strip().lower()
    if normalized in labels:
        return labels[normalized]
    return " ".join(part.capitalize() for part in normalized.split("_") if part) or "Metric"


def _snapshot_date(start_date: datetime | None, end_date: datetime | None) -> datetime | None:
    return end_date or start_date


def _insight_reporting_date(insight: SocialInsight) -> datetime | None:
    date_value = _snapshot_date(insight.start_date, insight.end_date)
    if (
        date_value
        and insight.platform == "facebook"
        and str(insight.period or "").lower() == "day"
        and insight.end_date is not None
    ):
        return date_value - timedelta(days=1)
    return date_value


def _metadata_source(metadata: dict[str, Any] | None) -> str | None:
    if not isinstance(metadata, dict):
        return None
    value = metadata.get("source")
    return str(value)[:80] if value else None


def _metadata_raw_metric(metric_name: str, metadata: dict[str, Any] | None) -> str:
    if not isinstance(metadata, dict):
        return metric_name
    for key in ("graph_metric", "raw_metric_name", "field"):
        value = metadata.get(key)
        if value:
            return str(value)[:160]
    candidate_metrics = metadata.get("candidate_metrics")
    if isinstance(candidate_metrics, list) and candidate_metrics:
        return str(candidate_metrics[0])[:160]
    candidate_fields = metadata.get("candidate_fields")
    if isinstance(candidate_fields, list) and candidate_fields:
        return str(candidate_fields[0])[:160]
    return metric_name


def _is_optional_daily_metric(platform: str, metric_name: str) -> bool:
    return metric_name in OPTIONAL_DAILY_METRICS.get(platform, set())


def _is_zero_when_empty_daily_metric(platform: str, metric_name: str) -> bool:
    return metric_name in ZERO_WHEN_EMPTY_DAILY_METRICS.get(platform, set())


def _is_permission_error(message: str) -> bool:
    text = message.lower()
    return "permission" in text or '"code": 10' in text or '"code":10' in text


def _is_metric_unavailable_error(message: str) -> bool:
    text = message.lower()
    return (
        "valid insights metric" in text
        or "must be a valid" in text
        or '"code": 100' in text
        or '"code":100' in text
        or "(#100)" in text
    )


TRANSIENT_GRAPH_STATUSES = {429, 500, 502, 503, 504}
TRANSIENT_GRAPH_CODES = {1, 2, 4, 17, 32, 613}


def _graph_log_params(params: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in params.items() if key != "access_token"}


def _redact_access_tokens(text: str) -> str:
    marker = "access_token="
    redacted = text
    search_from = 0
    while True:
        index = redacted.find(marker, search_from)
        if index < 0:
            return redacted
        value_start = index + len(marker)
        value_end = value_start
        while value_end < len(redacted) and redacted[value_end] not in {"&", '"', "'", " ", "\n", "\r", "\\", "}"}:
            value_end += 1
        redacted = f"{redacted[:value_start]}[REDACTED]{redacted[value_end:]}"
        search_from = value_start + len("[REDACTED]")


def _graph_log_payload(payload: Any) -> str:
    try:
        return _redact_access_tokens(json.dumps(payload))[:1500]
    except Exception:
        return _redact_access_tokens(str(payload))[:1500]


def _graph_log_fields(params: dict[str, Any]) -> dict[str, Any]:
    return {
        "metric": params.get("metric"),
        "period": params.get("period"),
        "metric_type": params.get("metric_type"),
        "since": params.get("since"),
        "until": params.get("until"),
    }


def _requires_total_value_retry(status_code: int, payload: Any, params: dict[str, Any]) -> bool:
    if status_code != 400 or params.get("metric_type"):
        return False
    metric = str(params.get("metric") or "").strip()
    if not metric:
        return False
    try:
        text = json.dumps(payload).lower()
    except Exception:
        text = str(payload).lower()
    return "should be specified with parameter metric_type=total_value" in text


def _graph_error_code(payload: Any) -> int | None:
    if not isinstance(payload, dict):
        return None
    error = payload.get("error")
    if not isinstance(error, dict):
        return None
    try:
        return int(error.get("code"))
    except (TypeError, ValueError):
        return None


def _graph_error_message(payload: Any) -> str:
    if not isinstance(payload, dict):
        return str(payload or "")
    error = payload.get("error")
    if isinstance(error, dict):
        return str(error.get("message") or "")
    return str(error or payload or "")


def _is_graph_unknown_path_response(status_code: int, payload: Any) -> bool:
    if status_code != 400:
        return False
    if _graph_error_code(payload) == 2500:
        return True
    return "unknown path components" in _graph_error_message(payload).lower()


def _is_graph_version_unavailable_response(status_code: int, payload: Any) -> bool:
    if status_code != 400:
        return False
    message = _graph_error_message(payload).lower()
    return (
        "unsupported version" in message
        or "unknown version" in message
        or "invalid version" in message
        or ("version" in message and "does not exist" in message)
        or ("version" in message and "not available" in message)
    )


def _graph_api_versions() -> list[str]:
    now = time.time()
    versions: list[str] = []
    for version in (SOCIAL_INSIGHTS_META_API_VERSION, SOCIAL_INSIGHTS_META_API_FALLBACK_VERSION):
        normalized = str(version or "").strip()
        if GRAPH_VERSION_UNAVAILABLE_UNTIL.get(normalized, 0) > now:
            continue
        if normalized and normalized not in versions:
            versions.append(normalized)
    return versions or ["v26.0"]


def _should_retry_graph_response(status_code: int, payload: Any) -> bool:
    if status_code in TRANSIENT_GRAPH_STATUSES:
        return True
    return _graph_error_code(payload) in TRANSIENT_GRAPH_CODES


def _graph_get(path: str, access_token: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    base_query = dict(params or {})
    base_query["access_token"] = access_token
    max_attempts = 3
    last_error: Exception | None = None
    versions = _graph_api_versions()
    for version_index, api_version in enumerate(versions):
        query = dict(base_query)
        url = f"https://graph.facebook.com/{api_version}/{path.lstrip('/')}"
        for attempt in range(1, max_attempts + 1):
            log_params = _graph_log_params(query)
            log_fields = _graph_log_fields(log_params)
            try:
                response = requests.get(url, params=query, timeout=API_TIMEOUT_SECONDS)
            except requests.RequestException as error:
                last_error = error
                logger.warning(
                    "Meta Graph request exception api_version=%s endpoint=%s metric=%s period=%s metric_type=%s since=%s until=%s attempt=%s/%s error=%s",
                    api_version,
                    path,
                    log_fields["metric"],
                    log_fields["period"],
                    log_fields["metric_type"],
                    log_fields["since"],
                    log_fields["until"],
                    attempt,
                    max_attempts,
                    _safe_error(error),
                )
                if attempt < max_attempts:
                    time.sleep(0.75 * attempt)
                    continue
                raise RuntimeError(_safe_error(error)) from error

            try:
                payload = response.json()
            except ValueError:
                payload = {"error": response.text[:500]}

            if response.ok:
                if isinstance(payload, dict):
                    payload.setdefault("_graph_api_version", api_version)
                logger.info(
                    "Meta Graph request ok api_version=%s endpoint=%s metric=%s period=%s metric_type=%s since=%s until=%s status=%s response=%s",
                    api_version,
                    path,
                    log_fields["metric"],
                    log_fields["period"],
                    log_fields["metric_type"],
                    log_fields["since"],
                    log_fields["until"],
                    response.status_code,
                    _graph_log_payload(payload),
                )
                return payload if isinstance(payload, dict) else {}

            if _requires_total_value_retry(response.status_code, payload, query):
                metric = str(query.get("metric") or "").strip()
                query["metric_type"] = "total_value"
                DISCOVERED_TOTAL_VALUE_INSIGHT_METRICS.add(metric)
                logger.warning(
                    "Meta Graph requested metric_type=total_value; retrying api_version=%s endpoint=%s metric=%s period=%s metric_type=total_value since=%s until=%s status=%s response=%s",
                    api_version,
                    path,
                    metric,
                    query.get("period"),
                    query.get("since"),
                    query.get("until"),
                    response.status_code,
                    _graph_log_payload(payload),
                )
                continue

            if (
                _is_graph_unknown_path_response(response.status_code, payload)
                or _is_graph_version_unavailable_response(response.status_code, payload)
            ) and version_index < len(versions) - 1:
                GRAPH_VERSION_UNAVAILABLE_UNTIL[api_version] = time.time() + GRAPH_VERSION_FALLBACK_CACHE_SECONDS
                logger.warning(
                    "Meta Graph path or version unavailable on api_version=%s; trying fallback endpoint=%s metric=%s period=%s metric_type=%s since=%s until=%s status=%s response=%s",
                    api_version,
                    path,
                    log_fields["metric"],
                    log_fields["period"],
                    log_fields["metric_type"],
                    log_fields["since"],
                    log_fields["until"],
                    response.status_code,
                    _graph_log_payload(payload),
                )
                break

            if attempt < max_attempts and _should_retry_graph_response(response.status_code, payload):
                logger.warning(
                    "Meta Graph request failed; retrying api_version=%s endpoint=%s metric=%s period=%s metric_type=%s since=%s until=%s status=%s attempt=%s/%s response=%s",
                    api_version,
                    path,
                    log_fields["metric"],
                    log_fields["period"],
                    log_fields["metric_type"],
                    log_fields["since"],
                    log_fields["until"],
                    response.status_code,
                    attempt,
                    max_attempts,
                    _graph_log_payload(payload),
                )
                time.sleep(0.75 * attempt)
                continue

            logger.warning(
                "Meta Graph request failed api_version=%s endpoint=%s metric=%s period=%s metric_type=%s since=%s until=%s status=%s response=%s",
                api_version,
                path,
                log_fields["metric"],
                log_fields["period"],
                log_fields["metric_type"],
                log_fields["since"],
                log_fields["until"],
                response.status_code,
                _graph_log_payload(payload),
            )
            error = payload.get("error") if isinstance(payload, dict) else payload
            raise RuntimeError(json.dumps(error))

    raise RuntimeError(_safe_error(last_error or "Meta API request failed."))


def _metric_number(value: Any) -> float | None:
    if isinstance(value, dict):
        if not value:
            return 0.0
        numbers = [_metric_number(item) for item in value.values()]
        cleaned = [item for item in numbers if item is not None]
        return sum(cleaned) if cleaned else None
    if isinstance(value, list):
        if not value:
            return 0.0
        numbers = [_metric_number(item) for item in value]
        cleaned = [item for item in numbers if item is not None]
        return sum(cleaned) if cleaned else None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _metadata_with_graph_metric(source: str, graph_metric: str, extra: dict[str, Any] | None = None) -> dict[str, Any]:
    metadata = {"source": source, "graph_metric": graph_metric}
    if extra:
        metadata.update(extra)
    return metadata


def _is_removed_metric_name(metric_name: str | None) -> bool:
    normalized = str(metric_name or "").strip().lower()
    return normalized in REMOVED_INSIGHT_METRICS or normalized.startswith("page_fan")


def _visible_insight_filter() -> Any:
    return (
        SocialInsight.metric_value.isnot(None),
        SocialInsight.error_message.is_(None),
        ~func.lower(SocialInsight.metric_name).in_(REMOVED_INSIGHT_METRICS),
        ~func.lower(SocialInsight.metric_name).like("page_fan%"),
    )


def _new_run_context() -> InsightRunContext:
    started_at = utcnow()
    return {"id": uuid.uuid4().hex, "started_at": started_at}


def _run_id(run: InsightRunContext | None) -> str | None:
    return str(run["id"]) if run and run.get("id") else None


def _run_started_at(run: InsightRunContext | None) -> datetime | None:
    return run.get("started_at") if run else None


def _latest_run_id_for_account(account_id: int) -> str | None:
    row = (
        db.session.query(SocialInsight.refresh_run_id)
        .filter(SocialInsight.social_account_id == account_id)
        .filter(SocialInsight.refresh_run_id.isnot(None))
        .filter(*_visible_insight_filter())
        .order_by(SocialInsight.refresh_run_started_at.desc(), SocialInsight.refreshed_at.desc())
        .first()
    )
    if not row:
        return None
    return row[0]


def _latest_refresh_run_id_for_account(account_id: int) -> str | None:
    row = (
        db.session.query(SocialInsight.refresh_run_id)
        .filter(SocialInsight.social_account_id == account_id)
        .filter(SocialInsight.refresh_run_id.isnot(None))
        .order_by(SocialInsight.refresh_run_started_at.desc(), SocialInsight.refreshed_at.desc())
        .first()
    )
    if not row:
        return None
    return row[0]


def _latest_legacy_refreshed_at_for_account(account_id: int) -> datetime | None:
    return (
        db.session.query(func.max(SocialInsight.refreshed_at))
        .filter(SocialInsight.social_account_id == account_id)
        .scalar()
    )


def _latest_visible_insights_for_account(account_id: int, *, limit: int | None = None) -> list[SocialInsight]:
    query = SocialInsight.query.filter(SocialInsight.social_account_id == account_id)
    latest_run_id = _latest_run_id_for_account(account_id)
    if latest_run_id:
        query = query.filter(SocialInsight.refresh_run_id == latest_run_id)
    else:
        latest_refreshed_at = _latest_legacy_refreshed_at_for_account(account_id)
        if not latest_refreshed_at:
            return []
        query = query.filter(SocialInsight.refreshed_at >= latest_refreshed_at - timedelta(minutes=15))

    query = query.filter(*_visible_insight_filter())
    query = query.order_by(SocialInsight.end_date.asc().nullsfirst(), SocialInsight.metric_name.asc())
    if limit:
        query = query.limit(limit)
    return query.all()


def _visible_insight_history_for_account(account_id: int, *, limit: int | None = None) -> list[SocialInsight]:
    query = SocialInsight.query.filter(SocialInsight.social_account_id == account_id)
    query = query.filter(*_visible_insight_filter())
    query = query.order_by(
        SocialInsight.end_date.asc().nullsfirst(),
        SocialInsight.start_date.asc().nullsfirst(),
        SocialInsight.refreshed_at.asc(),
        SocialInsight.metric_name.asc(),
    )
    if limit:
        query = query.limit(limit)
    return query.all()


def _diagnostic_insights_for_account(account_id: int, *, limit: int = 80) -> list[SocialInsight]:
    query = SocialInsight.query.filter(SocialInsight.social_account_id == account_id)
    latest_run_id = _latest_refresh_run_id_for_account(account_id)
    if latest_run_id:
        query = query.filter(SocialInsight.refresh_run_id == latest_run_id)
    else:
        latest_refreshed_at = _latest_legacy_refreshed_at_for_account(account_id)
        if not latest_refreshed_at:
            return []
        query = query.filter(SocialInsight.refreshed_at >= latest_refreshed_at - timedelta(minutes=15))

    rows = (
        query.filter(
            or_(
                SocialInsight.error_message.isnot(None),
                SocialInsight.metric_value.is_(None),
            )
        )
        .filter(~func.lower(SocialInsight.metric_name).in_(REMOVED_INSIGHT_METRICS))
        .filter(~func.lower(SocialInsight.metric_name).like("page_fan%"))
        .filter(
            ~(
                (func.lower(SocialInsight.metric_name) == "reactions")
                & SocialInsight.error_message.is_(None)
            )
        )
        .order_by(SocialInsight.refreshed_at.desc(), SocialInsight.metric_name.asc())
        .limit(limit)
        .all()
    )
    return rows


def _cleanup_removed_insight_metrics() -> None:
    removed = (
        SocialInsight.query.filter(
            or_(
                func.lower(SocialInsight.metric_name).in_(REMOVED_INSIGHT_METRICS),
                func.lower(SocialInsight.metric_name).like("page_fan%"),
            )
        ).delete(synchronize_session=False)
    )
    removed_snapshots = (
        AccountInsightSnapshot.query.filter(
            or_(
                func.lower(AccountInsightSnapshot.metric_name).in_(REMOVED_INSIGHT_METRICS),
                func.lower(AccountInsightSnapshot.raw_metric_name).in_(REMOVED_INSIGHT_METRICS),
                func.lower(AccountInsightSnapshot.metric_name).like("page_fan%"),
                func.lower(AccountInsightSnapshot.raw_metric_name).like("page_fan%"),
            )
        ).delete(synchronize_session=False)
    )
    if removed or removed_snapshots:
        db.session.commit()


def _cleanup_instagram_legacy_follower_diagnostics(account: SocialAccount) -> None:
    if account.platform != "instagram" or not account.id:
        return
    SocialInsight.query.filter_by(social_account_id=account.id, metric_name="followers").filter(
        SocialInsight.metric_value.is_(None)
    ).delete(synchronize_session=False)


def _ensure_meta_token(account: SocialAccount) -> None:
    if account.access_token:
        return
    try:
        from ..integrations import apply_global_meta_token_to_account
        from ..settings import global_meta_user_token
    except Exception:
        return

    if not global_meta_user_token():
        return
    apply_global_meta_token_to_account(account)


def _mirror_account_snapshot(
    account: SocialAccount,
    *,
    metric_name: str,
    metric_value: float | None,
    period: str,
    start_date: datetime | None,
    end_date: datetime | None,
    metadata: dict[str, Any] | None = None,
    error_message: str | None = None,
) -> AccountInsightSnapshot:
    raw_metric_name = _metadata_raw_metric(metric_name, metadata)
    metadata_status = metadata.get("status") if isinstance(metadata, dict) else None
    status = (
        "error"
        if error_message
        else str(metadata_status)[:40]
        if metric_value is None and metadata_status
        else "unavailable"
        if metric_value is None
        else "ok"
    )
    date_value = _snapshot_date(start_date, end_date)
    snapshot = AccountInsightSnapshot.query.filter_by(
        social_account_id=account.id,
        metric_name=metric_name,
        period=period,
        date=date_value,
        raw_metric_name=raw_metric_name,
    ).first()
    if not snapshot:
        snapshot = AccountInsightSnapshot(
            social_account_id=account.id,
            metric_name=metric_name,
            period=period,
            date=date_value,
            raw_metric_name=raw_metric_name,
        )
        db.session.add(snapshot)

    changed = (
        snapshot.platform != account.platform
        or snapshot.metric_value != metric_value
        or snapshot.source != _metadata_source(metadata)
        or snapshot.status != status
        or snapshot.error_message != error_message
    )
    snapshot.platform = account.platform
    snapshot.metric_value = metric_value
    snapshot.source = _metadata_source(metadata)
    snapshot.status = status
    snapshot.error_message = error_message
    if changed or not snapshot.fetched_at:
        snapshot.fetched_at = utcnow()
    return snapshot


def _cleanup_daily_account_snapshot_duplicates(
    account: SocialAccount,
    *,
    metric_name: str,
    day: datetime,
    keep_insight_id: int | None,
) -> None:
    day_end = day + timedelta(days=1)
    insight_query = (
        SocialInsight.query.filter_by(
            social_account_id=account.id,
            metric_name=metric_name,
            period="snapshot",
        )
        .filter(SocialInsight.start_date >= day)
        .filter(SocialInsight.start_date < day_end)
    )
    if keep_insight_id:
        insight_query = insight_query.filter(SocialInsight.id != keep_insight_id)
    insight_query.delete(synchronize_session=False)

    keep_snapshot = AccountInsightSnapshot.query.filter_by(
        social_account_id=account.id,
        metric_name=metric_name,
        period="snapshot",
        date=day,
    ).first()
    snapshot_query = (
        AccountInsightSnapshot.query.filter_by(
            social_account_id=account.id,
            metric_name=metric_name,
            period="snapshot",
        )
        .filter(AccountInsightSnapshot.date >= day)
        .filter(AccountInsightSnapshot.date < day_end)
    )
    if keep_snapshot:
        snapshot_query = snapshot_query.filter(AccountInsightSnapshot.id != keep_snapshot.id)
    snapshot_query.delete(synchronize_session=False)


def _upsert_insight(
    account: SocialAccount,
    *,
    metric_name: str,
    metric_value: float | None,
    period: str,
    start_date: datetime | None,
    end_date: datetime | None,
    metadata: dict[str, Any] | None = None,
    error_message: str | None = None,
    run: InsightRunContext | None = None,
) -> SocialInsight:
    if _is_removed_metric_name(metric_name):
        raise RuntimeError("This insight metric is no longer supported.")

    insight = SocialInsight.query.filter_by(
        social_account_id=account.id,
        metric_name=metric_name,
        period=period,
        start_date=start_date,
        end_date=end_date,
    ).first()
    if not insight:
        insight = SocialInsight(
            social_account_id=account.id,
            platform=account.platform,
            metric_name=metric_name,
            period=period,
            start_date=start_date,
            end_date=end_date,
        )
        db.session.add(insight)

    insight.platform = account.platform
    insight.metric_value = metric_value
    insight.source_metadata = json.dumps(metadata or {})
    insight.refreshed_at = utcnow()
    insight.refresh_run_id = _run_id(run)
    insight.refresh_run_started_at = _run_started_at(run)
    if error_message:
        insight.last_error_at = insight.refreshed_at
        insight.error_message = error_message
    else:
        insight.last_success_at = insight.refreshed_at
        insight.error_message = None
    _mirror_account_snapshot(
        account,
        metric_name=metric_name,
        metric_value=metric_value,
        period=period,
        start_date=start_date,
        end_date=end_date,
        metadata=metadata,
        error_message=error_message,
    )
    return insight


def _store_account_error(account: SocialAccount, message: str, run: InsightRunContext | None = None) -> None:
    now = utcnow()
    for metric_name in _metric_names_for_platform(account.platform) or ["account_insights"]:
        _upsert_insight(
            account,
            metric_name=metric_name,
            metric_value=None,
            period="error",
            start_date=None,
            end_date=None,
            metadata={"platform": account.platform},
            error_message=message,
            run=run,
        )
    account.test_status = "failed"
    account.test_error = message
    account.last_tested = now


def _store_metric_error(
    account: SocialAccount,
    *,
    metric_name: str,
    period: str,
    start_date: datetime | None,
    end_date: datetime | None,
    message: str,
    metadata: dict[str, Any] | None = None,
    run: InsightRunContext | None = None,
) -> SocialInsight:
    return _upsert_insight(
        account,
        metric_name=metric_name,
        metric_value=None,
        period=period,
        start_date=start_date,
        end_date=end_date,
        metadata=metadata,
        error_message=_safe_error(message),
        run=run,
    )


def _store_metric_unavailable(
    account: SocialAccount,
    *,
    metric_name: str,
    period: str,
    start_date: datetime | None,
    end_date: datetime | None,
    metadata: dict[str, Any] | None = None,
    run: InsightRunContext | None = None,
) -> SocialInsight:
    details = dict(metadata or {})
    details["availability"] = "unavailable"
    return _upsert_insight(
        account,
        metric_name=metric_name,
        metric_value=None,
        period=period,
        start_date=start_date,
        end_date=end_date,
        metadata=details,
        error_message=None,
        run=run,
    )


def _store_instagram_follower_snapshot(
    account: SocialAccount,
    *,
    payload: dict[str, Any],
    followers_count: float,
    captured_at: datetime,
) -> InstagramFollowerSnapshot | None:
    if account.platform != "instagram" or not account.page_id_external:
        return None
    ig_user_id = str(payload.get("id") or account.page_id_external)
    day_start = captured_at.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = day_start + timedelta(days=1)
    existing = (
        InstagramFollowerSnapshot.query.filter_by(account_id=account.id, ig_user_id=ig_user_id)
        .filter(InstagramFollowerSnapshot.captured_at >= day_start)
        .filter(InstagramFollowerSnapshot.captured_at < day_end)
        .order_by(InstagramFollowerSnapshot.captured_at.desc(), InstagramFollowerSnapshot.id.desc())
        .all()
    )
    if existing:
        snapshot = existing[0]
        for duplicate in existing[1:]:
            db.session.delete(duplicate)
    else:
        snapshot = InstagramFollowerSnapshot(account_id=account.id, ig_user_id=ig_user_id)
        db.session.add(snapshot)
    snapshot.username = str(payload.get("username") or account.account_name or "")[:160] or None
    snapshot.followers_count = int(round(followers_count))
    snapshot.captured_at = captured_at
    return snapshot


def _refresh_field_metric_group(
    account: SocialAccount,
    metric_name: str,
    fields: list[str],
    run: InsightRunContext | None = None,
) -> list[SocialInsight]:
    if not fields or not account.page_id_external or not account.access_token:
        return []

    errors: list[str] = []
    for field in fields:
        try:
            payload = _graph_get(
                account.page_id_external,
                account.access_token,
                {"fields": field},
            )
        except Exception as error:
            errors.append(f"{field}: {_safe_error(error)}")
            continue

        metric_value = _metric_number(payload.get(field))
        if metric_value is None:
            errors.append(f"{field}: no numeric value returned")
            continue

        return [
            _upsert_insight(
                account,
                metric_name=metric_name,
                metric_value=metric_value,
                period="lifetime",
                start_date=None,
                end_date=None,
                metadata=_metadata_with_graph_metric("fields", field),
                run=run,
            )
        ]

    if errors:
        return [
            _store_metric_error(
                account,
                metric_name=metric_name,
                period="lifetime",
                start_date=None,
                end_date=None,
                message="; ".join(errors),
                metadata={"source": "fields", "candidate_fields": fields},
                run=run,
            )
        ]
    return []


def _refresh_field_metrics(account: SocialAccount, run: InsightRunContext | None = None) -> list[SocialInsight]:
    if not account.page_id_external or not account.access_token:
        return []

    fields = ACCOUNT_OBJECT_FIELDS.get(account.platform, [])
    field_metrics = ACCOUNT_OBJECT_FIELD_METRICS.get(account.platform, {})
    if not fields or not field_metrics:
        return []

    payload: dict[str, Any] = {}
    try:
        payload = _graph_get(account.page_id_external, account.access_token, {"fields": ",".join(fields)})
    except Exception as combined_error:
        logger.info(
            "Combined object field lookup failed for account_id=%s platform=%s: %s",
            account.id,
            account.platform,
            _safe_error(combined_error),
        )
        for field in fields:
            try:
                payload.update(_graph_get(account.page_id_external, account.access_token, {"fields": field}))
            except Exception:
                continue

    display_name = payload.get("username") or payload.get("name")
    if display_name:
        account.account_name = str(display_name)[:100]

    stored: list[SocialInsight] = []
    seen_metrics: set[str] = set()
    missing_by_metric: dict[str, list[str]] = {}
    captured_at = utcnow()
    captured_day = captured_at.replace(hour=0, minute=0, second=0, microsecond=0)
    for field in fields:
        metric_name = field_metrics.get(field)
        if not metric_name or metric_name in seen_metrics:
            continue
        metric_value = _metric_number(payload.get(field))
        if metric_value is None:
            missing_by_metric.setdefault(metric_name, []).append(field)
            continue
        seen_metrics.add(metric_name)
        if account.platform == "instagram" and field == "followers_count":
            _store_instagram_follower_snapshot(
                account,
                payload=payload,
                followers_count=metric_value,
                captured_at=captured_at,
            )
        is_daily_snapshot = account.platform == "instagram" and field in IG_USER_FIELD_METRICS
        insight = _upsert_insight(
            account,
            metric_name=metric_name,
            metric_value=metric_value,
            period="snapshot" if is_daily_snapshot else "lifetime",
            start_date=captured_day if is_daily_snapshot else None,
            end_date=captured_day if is_daily_snapshot else None,
            metadata=_metadata_with_graph_metric(
                "fields",
                field,
                {
                    "graph_version": SOCIAL_INSIGHTS_META_API_VERSION,
                    "endpoint": account.page_id_external,
                    "request_params": {"fields": ",".join(fields)},
                    "response_shape": "ig_user_field" if is_daily_snapshot else "object_field",
                },
            ),
            run=run,
        )
        if is_daily_snapshot:
            db.session.flush()
            _cleanup_daily_account_snapshot_duplicates(
                account,
                metric_name=metric_name,
                day=captured_day,
                keep_insight_id=insight.id,
            )
        stored.append(insight)

    for metric_name, missing_fields in missing_by_metric.items():
        if metric_name in seen_metrics:
            continue
        stored.append(
            _store_metric_unavailable(
                account,
                metric_name=metric_name,
                period="lifetime",
                start_date=None,
                end_date=None,
                metadata={"source": "fields", "candidate_fields": missing_fields, "reason": "no_data"},
                run=run,
            )
        )
    return stored


def _refresh_daily_metric(
    account: SocialAccount,
    metric_name: str,
    candidate: InsightMetricCandidate,
    start: datetime,
    end: datetime,
    run: InsightRunContext | None = None,
) -> list[SocialInsight]:
    if not account.page_id_external or not account.access_token:
        raise RuntimeError("Account is missing Meta object ID or access token.")
    graph_metric = str(candidate.get("metric") or "").strip()
    if not graph_metric:
        return []
    params = _account_insight_candidate_params(account.platform, graph_metric, candidate)
    params["metric"] = graph_metric
    if str(params.get("period") or "").lower() != "lifetime":
        max_days = int(candidate.get("max_days") or (30 if account.platform == "instagram" else 90))
        windows = _insight_date_windows(start, end, max_days=max_days)
        if len(windows) > 1:
            stored: list[SocialInsight] = []
            for chunk_start, chunk_end in windows:
                stored.extend(_refresh_daily_metric(account, metric_name, candidate, chunk_start, chunk_end, run))
            return stored
        params["since"] = start.strftime("%Y-%m-%d")
        params["until"] = _graph_until_date(end)
    payload = _graph_get(
        f"{account.page_id_external}/insights",
        account.access_token,
        params,
    )
    data = payload.get("data") if isinstance(payload, dict) else []
    if not isinstance(data, list):
        return []
    stored: list[SocialInsight] = []
    for metric in data:
        if not isinstance(metric, dict):
            continue

        total_value = metric.get("total_value")
        if total_value is not None:
            metric_value = _metric_number(total_value.get("value") if isinstance(total_value, dict) else total_value)
            if metric_value is not None:
                stored.append(
                    _upsert_insight(
                        account,
                        metric_name=metric_name,
                        metric_value=metric_value,
                        period=str(params.get("period") or "day"),
                        start_date=start,
                        end_date=end,
                        metadata=_metadata_with_graph_metric(
                            "insights",
                            graph_metric,
                            {
                                "graph_version": SOCIAL_INSIGHTS_META_API_VERSION,
                                "endpoint": f"{account.page_id_external}/insights",
                                "request_params": _graph_log_params(params),
                                "response_shape": "total_value",
                            },
                        ),
                        run=run,
                    )
                )
            continue

        values = metric.get("values")
        if not isinstance(values, list):
            continue
        for item in values:
            if not isinstance(item, dict):
                continue
            end_time = item.get("end_time")
            try:
                end_date = core.parse_iso_datetime(str(end_time)) if end_time else None
            except Exception:
                end_date = None
            value = item.get("value")
            metric_value = _metric_number(value)
            period = str(params.get("period") or "day")
            start_date = None
            stored_end_date = end_date
            response_shape = "daily_series"
            if account.platform == "facebook" and metric_name == "followers" and graph_metric == "page_follows" and end_date:
                snapshot_date = _start_of_day(end_date - timedelta(days=1))
                period = "snapshot"
                start_date = snapshot_date
                stored_end_date = snapshot_date
                response_shape = "daily_snapshot"
            stored.append(
                _upsert_insight(
                    account,
                    metric_name=metric_name,
                    metric_value=metric_value,
                    period=period,
                    start_date=start_date,
                    end_date=stored_end_date,
                    metadata=_metadata_with_graph_metric(
                        "insights",
                        graph_metric,
                        {
                            "graph_version": SOCIAL_INSIGHTS_META_API_VERSION,
                            "endpoint": f"{account.page_id_external}/insights",
                            "request_params": _graph_log_params(params),
                            "graph_end_time": end_time,
                            "response_shape": response_shape,
                        },
                    ),
                    run=run,
                )
            )
    return stored


def _refresh_daily_metric_group(
    account: SocialAccount,
    metric_name: str,
    graph_metrics: list[InsightMetricCandidate],
    start: datetime,
    end: datetime,
    run: InsightRunContext | None = None,
) -> list[SocialInsight]:
    errors: list[str] = []
    candidate_names = [str(candidate.get("metric") or "") for candidate in graph_metrics]
    candidate_requests = [
        {
            "metric": str(candidate.get("metric") or ""),
            "params": _account_insight_candidate_params(
                account.platform,
                str(candidate.get("metric") or "").strip(),
                candidate,
            ),
            "endpoint": f"{account.page_id_external}/insights",
            "graph_version": SOCIAL_INSIGHTS_META_API_VERSION,
        }
        for candidate in graph_metrics
    ]
    for candidate in graph_metrics:
        graph_metric = str(candidate.get("metric") or "").strip()
        candidate_params = dict(candidate.get("params") or {})
        candidate_label = f"{graph_metric} {candidate_params}".strip()
        try:
            stored = _refresh_daily_metric(account, metric_name, candidate, start, end, run)
        except Exception as error:
            errors.append(f"{candidate_label}: {_safe_error(error)}")
            continue

        if stored:
            return stored
        errors.append(f"{candidate_label}: no values returned")

    if _is_zero_when_empty_daily_metric(account.platform, metric_name) and all(
        "no values returned" in error for error in errors
    ):
        return [
            _upsert_insight(
                account,
                metric_name=metric_name,
                metric_value=0,
                period="day",
                start_date=start,
                end_date=end,
                metadata={
                    "source": "insights",
                    "candidate_metrics": candidate_names,
                    "candidate_requests": candidate_requests,
                    "note": "Meta returned no rows; stored as zero activity for the selected range.",
                },
                run=run,
            )
        ]

    if _is_optional_daily_metric(account.platform, metric_name) and (
        not errors
        or any(_is_permission_error(error) for error in errors)
        or any(_is_metric_unavailable_error(error) for error in errors)
        or all("no values returned" in error for error in errors)
    ):
        return [
            _store_metric_unavailable(
                account,
                metric_name=metric_name,
                period="day",
                start_date=start,
                end_date=end,
                metadata={
                    "source": "insights",
                    "candidate_metrics": candidate_names,
                    "candidate_requests": candidate_requests,
                    "reason": "no_data" if all("no values returned" in error for error in errors) else "permission_unavailable" if errors else "no_data",
                },
                run=run,
            )
        ]

    return [
        _store_metric_error(
            account,
            metric_name=metric_name,
            period="day",
            start_date=start,
            end_date=end,
            message="; ".join(errors) if errors else "No data returned.",
            metadata={"source": "insights", "candidate_metrics": candidate_names, "candidate_requests": candidate_requests},
            run=run,
        )
    ]


def _platform_post_id_for_post(post: Post, platform: str) -> str | None:
    value = {
        "facebook": post.facebook_post_id or post.facebook_remote_post_id,
        "instagram": post.instagram_post_id,
        "linkedin": post.linkedin_post_id,
        "twitter": post.twitter_post_id,
        "pinterest": post.pinterest_post_id,
    }.get(platform)
    return str(value).strip() if value else None


def _ensure_post_references_for_account(account: SocialAccount) -> list[PlatformPostReference]:
    if not account.id:
        return []
    posts = (
        Post.query.filter(Post.page_id == account.page_id)
        .filter(Post.status.in_(["posted", "manual_pending"]))
        .order_by(Post.posted_at.desc().nullslast(), Post.id.desc())
        .limit(1000)
        .all()
    )
    references: list[PlatformPostReference] = []
    for post in posts:
        platform_post_id = _platform_post_id_for_post(post, account.platform)
        if not platform_post_id:
            continue
        reference = PlatformPostReference.query.filter_by(
            internal_post_id=post.id,
            social_account_id=account.id,
            platform_post_id=platform_post_id,
        ).first()
        if not reference:
            reference = PlatformPostReference(
                internal_post_id=post.id,
                social_account_id=account.id,
                platform=account.platform,
                platform_post_id=platform_post_id,
            )
            db.session.add(reference)
        reference.platform = account.platform
        reference.permalink = post.platform_url_map().get(account.platform)
        reference.published_at = post.posted_at or post.scheduled_time or post.created_at
        reference.media_type = post.media_type
        reference.caption_preview = (post.content or "")[:280]
        references.append(reference)
    return references


def _upsert_post_insight(
    reference: PlatformPostReference,
    *,
    metric_name: str,
    metric_value: float | None,
    period: str,
    date_value: datetime | None,
    status: str,
    error_message: str | None = None,
) -> PostInsightSnapshot:
    base_query = PostInsightSnapshot.query.filter_by(
        internal_post_id=reference.internal_post_id,
        social_account_id=reference.social_account_id,
        platform_post_id=reference.platform_post_id,
        metric_name=metric_name,
        period=period,
    )
    snapshot = base_query.filter_by(
        date=date_value,
    ).first()
    if not snapshot and period == "lifetime":
        snapshot = base_query.order_by(PostInsightSnapshot.fetched_at.desc(), PostInsightSnapshot.id.desc()).first()
    if not snapshot:
        snapshot = PostInsightSnapshot(
            internal_post_id=reference.internal_post_id,
            social_account_id=reference.social_account_id,
            platform_post_id=reference.platform_post_id,
            metric_name=metric_name,
            period=period,
            date=date_value,
        )
        db.session.add(snapshot)
    elif period == "lifetime":
        for duplicate in base_query.filter(PostInsightSnapshot.id != snapshot.id).all():
            db.session.delete(duplicate)

    if metric_value is None and status == "unavailable":
        previous = (
            base_query.filter(PostInsightSnapshot.metric_value.isnot(None))
            .order_by(PostInsightSnapshot.fetched_at.desc().nullslast(), PostInsightSnapshot.id.desc())
            .first()
        )
        if previous:
            metric_value = previous.metric_value
            status = "stale"
            error_message = None

    changed = (
        snapshot.platform != reference.platform
        or snapshot.metric_value != metric_value
        or snapshot.status != status
        or snapshot.error_message != error_message
        or snapshot.date != date_value
    )
    snapshot.platform = reference.platform
    snapshot.metric_value = metric_value
    snapshot.status = status
    snapshot.error_message = error_message
    snapshot.date = date_value
    if changed or not snapshot.fetched_at:
        snapshot.fetched_at = utcnow()
    return snapshot


def _media_metric_period(metric: dict[str, Any]) -> str:
    return str(metric.get("period") or "lifetime").strip().lower() or "lifetime"


def _media_metric_value(metric: dict[str, Any], *, prefer_lifetime: bool = True) -> float | None:
    total_value = metric.get("total_value")
    if total_value is not None:
        return _metric_number(total_value.get("value") if isinstance(total_value, dict) else total_value)

    raw_values = metric.get("values")
    if not isinstance(raw_values, list):
        return None

    numbers: list[float] = []
    for item in raw_values:
        if not isinstance(item, dict):
            continue
        number = _metric_number(item.get("value"))
        if number is not None:
            numbers.append(number)
    if not numbers:
        return None
    if prefer_lifetime or _media_metric_period(metric) == "lifetime":
        return numbers[0]
    return sum(numbers)


def _media_insight_value(payload: dict[str, Any], *, prefer_lifetime: bool = True) -> float | None:
    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, list):
        return None
    lifetime_values: list[float] = []
    fallback_values: list[float] = []
    for metric in data:
        if not isinstance(metric, dict):
            continue
        number = _media_metric_value(metric, prefer_lifetime=prefer_lifetime)
        if number is None:
            continue
        if _media_metric_period(metric) == "lifetime":
            lifetime_values.append(number)
        else:
            fallback_values.append(number)
    values = lifetime_values or fallback_values
    return sum(values) if values else None


def _media_insight_values_by_name(payload: dict[str, Any], *, prefer_lifetime: bool = True) -> dict[str, float]:
    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, list):
        return {}
    values: dict[str, float] = {}
    periods: dict[str, str] = {}
    for metric in data:
        if not isinstance(metric, dict):
            continue
        name = str(metric.get("name") or metric.get("metric") or "").strip()
        if not name:
            continue
        number = _media_metric_value(metric, prefer_lifetime=prefer_lifetime)
        if number is not None:
            period = _media_metric_period(metric)
            if name in values and periods.get(name) == "lifetime" and period != "lifetime":
                continue
            values[name] = number
            periods[name] = period
    return values


def _parse_graph_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    text = str(value).strip()
    if len(text) >= 5 and text[-5] in {"+", "-"} and text[-3] != ":":
        text = f"{text[:-2]}:{text[-2:]}"
    try:
        return core.parse_iso_datetime(text)
    except Exception:
        return None


def _normalize_post_match_text(value: Any) -> str:
    return " ".join(str(value or "").lower().split())


def _captions_match_for_reference(local_caption: str | None, remote_caption: str | None) -> bool:
    local = _normalize_post_match_text(local_caption)
    remote = _normalize_post_match_text(remote_caption)
    if not local or not remote:
        return False
    if local == remote:
        return True
    shortest = min(len(local), len(remote))
    if shortest < 32:
        return False
    if local[:160] == remote[:160]:
        return True
    return shortest >= 80 and (local in remote or remote in local)


def _post_reference_match_window_hours(caption: str | None) -> int:
    return 48 if len(_normalize_post_match_text(caption)) >= 80 else 12


def _matching_internal_post_for_remote_post(
    account: SocialAccount,
    *,
    platform: str,
    platform_post_id: str,
    caption: str | None,
    published_at: datetime | None,
) -> Post | None:
    existing_reference = PlatformPostReference.query.filter_by(
        social_account_id=account.id,
        platform_post_id=platform_post_id,
    ).first()
    if existing_reference:
        return existing_reference.post

    platform_id_query = Post.query.filter(Post.page_id == account.page_id)
    if platform == "facebook":
        platform_id_query = platform_id_query.filter(
            or_(
                Post.facebook_post_id == platform_post_id,
                Post.facebook_remote_post_id == platform_post_id,
            )
        )
    elif platform == "instagram":
        platform_id_query = platform_id_query.filter(Post.instagram_post_id == platform_post_id)
    else:
        platform_id_query = platform_id_query.filter(False)
    platform_match = platform_id_query.first()
    if platform_match:
        return platform_match

    if not published_at or not caption:
        return None

    candidates = (
        Post.query.filter(Post.page_id == account.page_id)
        .filter(Post.status.in_(["posted", "manual_pending"]))
        .order_by(Post.posted_at.desc().nullslast(), Post.scheduled_time.desc().nullslast(), Post.id.desc())
        .limit(500)
        .all()
    )
    max_delta_seconds = _post_reference_match_window_hours(caption) * 3600
    best: tuple[float, Post] | None = None
    for post in candidates:
        existing_platform_id = _platform_post_id_for_post(post, platform)
        if existing_platform_id and existing_platform_id != platform_post_id:
            continue
        if not _captions_match_for_reference(post.content, caption):
            continue
        local_date = post.posted_at or post.scheduled_time or post.created_at
        if not local_date:
            continue
        delta_seconds = abs((published_at - local_date).total_seconds())
        if delta_seconds > max_delta_seconds:
            continue
        if not best or delta_seconds < best[0]:
            best = (delta_seconds, post)
    return best[1] if best else None


def _set_post_platform_reference_fields(post: Post, platform: str, platform_post_id: str, permalink: str | None) -> None:
    if platform == "facebook" and not post.facebook_post_id:
        post.facebook_post_id = platform_post_id
    elif platform == "instagram" and not post.instagram_post_id:
        post.instagram_post_id = platform_post_id

    platforms = post.platform_list()
    if platform not in platforms:
        platforms.append(platform)
        post.platforms = json.dumps(platforms)

    if permalink:
        url_map = post.platform_url_map()
        url_map[platform] = permalink
        post.platform_post_urls = json.dumps(url_map) if url_map else None


def _remote_media_type(platform: str, media_type: str | None) -> str | None:
    normalized = str(media_type or "").strip().lower()
    if not normalized:
        return None
    if platform == "instagram":
        if "video" in normalized or "reel" in normalized:
            return "video"
        if "carousel" in normalized:
            return "mixed"
        return "image"
    if "video" in normalized:
        return "video"
    if "photo" in normalized or "image" in normalized:
        return "image"
    return normalized[:20]


def _set_remote_post_media(post: Post, thumbnail: str | None) -> None:
    if not thumbnail:
        return
    media_items = post.media_list()
    if media_items:
        return
    post.media_paths = json.dumps([str(thumbnail)])


def _create_remote_post_for_reference(
    account: SocialAccount,
    *,
    platform: str,
    platform_post_id: str,
    permalink: str | None,
    published_at: datetime | None,
    media_type: str | None,
    caption: str | None,
    thumbnail: str | None,
) -> Post:
    timestamp = published_at or utcnow()
    post = Post(
        page_id=account.page_id,
        content=str(caption or "")[:10000],
        media_type=_remote_media_type(platform, media_type),
        platforms=json.dumps([platform]),
        scheduled_time=published_at,
        status="posted",
        created_at=timestamp,
        posted_at=published_at,
    )
    _set_remote_post_media(post, thumbnail)
    db.session.add(post)
    db.session.flush()
    _set_post_platform_reference_fields(post, platform, platform_post_id, permalink)
    return post


def _upsert_remote_platform_reference(
    account: SocialAccount,
    *,
    platform: str,
    platform_post_id: str,
    permalink: str | None,
    published_at: datetime | None,
    media_type: str | None,
    caption: str | None,
    thumbnail: str | None = None,
) -> PlatformPostReference | None:
    post = _matching_internal_post_for_remote_post(
        account,
        platform=platform,
        platform_post_id=platform_post_id,
        caption=caption,
        published_at=published_at,
    )
    if not post:
        post = _create_remote_post_for_reference(
            account,
            platform=platform,
            platform_post_id=platform_post_id,
            permalink=permalink,
            published_at=published_at,
            media_type=media_type,
            caption=caption,
            thumbnail=thumbnail,
        )

    reference = PlatformPostReference.query.filter_by(
        internal_post_id=post.id,
        social_account_id=account.id,
        platform_post_id=platform_post_id,
    ).first()
    if not reference:
        reference = PlatformPostReference(
            internal_post_id=post.id,
            social_account_id=account.id,
            platform=platform,
            platform_post_id=platform_post_id,
        )
        db.session.add(reference)

    reference.platform = platform
    reference.permalink = permalink or post.platform_url_map().get(platform)
    reference.published_at = published_at or post.posted_at or post.scheduled_time or post.created_at
    reference.media_type = media_type
    reference.caption_preview = (caption or post.content or "")[:280]
    _set_remote_post_media(post, thumbnail)
    _set_post_platform_reference_fields(post, platform, platform_post_id, reference.permalink)
    return reference


def _store_post_metric_values(reference: PlatformPostReference, metrics: dict[str, Any]) -> list[PostInsightSnapshot]:
    stored: list[PostInsightSnapshot] = []
    allowed_metrics = set(POST_INSIGHT_METRICS.get(reference.platform, []))
    for metric_name, metric_value in metrics.items():
        if metric_name not in allowed_metrics:
            continue
        stored.append(
            _upsert_post_insight(
                reference,
                metric_name=metric_name,
                metric_value=float(metric_value) if metric_value is not None else None,
                period="lifetime",
                date_value=reference.published_at,
                status="ok" if metric_value is not None else "unavailable",
            )
        )
    return stored


def _summary_total_count(value: Any) -> float | None:
    if not isinstance(value, dict):
        return None
    summary = value.get("summary")
    if isinstance(summary, dict):
        return _metric_number(summary.get("total_count"))
    return None


def _facebook_post_field_metrics(payload: dict[str, Any]) -> dict[str, float]:
    metrics: dict[str, float] = {}
    shares = payload.get("shares") if isinstance(payload, dict) else None
    if isinstance(shares, dict):
        value = _metric_number(shares.get("count"))
        if value is not None:
            metrics["shares"] = value
    elif isinstance(payload, dict):
        metrics["shares"] = 0.0

    comments = _summary_total_count(payload.get("comments")) if isinstance(payload, dict) else None
    if comments is not None:
        metrics["comments"] = comments

    reactions = _summary_total_count(payload.get("reactions")) if isinstance(payload, dict) else None
    if reactions is not None:
        metrics["reactions"] = reactions

    if any(key in metrics for key in ("comments", "shares", "reactions")):
        metrics.setdefault(
            "engagement",
            sum(float(metrics.get(key) or 0) for key in ("comments", "shares", "reactions")),
        )
    return metrics


def _facebook_post_thumbnail(payload: dict[str, Any]) -> str | None:
    full_picture = payload.get("full_picture") if isinstance(payload, dict) else None
    if full_picture:
        return str(full_picture)
    attachments = payload.get("attachments") if isinstance(payload, dict) else None
    data = attachments.get("data") if isinstance(attachments, dict) else None
    if not isinstance(data, list):
        return None
    for attachment in data:
        if not isinstance(attachment, dict):
            continue
        media = attachment.get("media")
        image = media.get("image") if isinstance(media, dict) else None
        source = image.get("src") if isinstance(image, dict) else None
        if source:
            return str(source)
    return None


def _fetch_facebook_post_insight_metrics(account: SocialAccount, post_id: str) -> dict[str, float]:
    values: dict[str, float] = {}
    for metric_name, graph_metrics in FACEBOOK_POST_INSIGHT_METRIC_PRIORITY.items():
        found_value = False
        for graph_metric in graph_metrics:
            seen_params: set[tuple[tuple[str, str], ...]] = set()
            for extra_params in FACEBOOK_POST_INSIGHT_PARAM_OPTIONS:
                params = {"metric": graph_metric, **extra_params}
                dedupe_key = tuple(sorted((str(key), str(value)) for key, value in params.items()))
                if dedupe_key in seen_params:
                    continue
                seen_params.add(dedupe_key)
                try:
                    payload = _graph_get(
                        f"{post_id}/insights",
                        account.access_token or "",
                        params,
                    )
                except Exception as error:
                    message = _safe_error(error)
                    if _is_metric_unavailable_error(message):
                        logger.debug(
                            "Facebook post insight metric unavailable account_id=%s post_id=%s metric=%s period=%s error=%s",
                            account.id,
                            post_id,
                            graph_metric,
                            params.get("period"),
                            message,
                        )
                    else:
                        logger.info(
                            "Facebook post insight unavailable account_id=%s post_id=%s metric=%s period=%s error=%s",
                            account.id,
                            post_id,
                            graph_metric,
                            params.get("period"),
                            message,
                        )
                    continue
                raw_values = _media_insight_values_by_name(payload)
                raw_value = raw_values.get(graph_metric)
                metric_value = raw_value if raw_value is not None else _media_insight_value(payload)
                if metric_value is not None:
                    values[metric_name] = metric_value
                    found_value = True
                    break
            if found_value:
                break
    return values


def _fetch_facebook_post_metrics(account: SocialAccount, post_id: str) -> dict[str, float]:
    if not account.access_token:
        return {}

    metrics: dict[str, float] = {}
    try:
        payload = _graph_get(post_id, account.access_token, {"fields": FACEBOOK_POST_FIELDS})
        metrics.update(_facebook_post_field_metrics(payload))
    except Exception as error:
        logger.info(
            "Facebook post fields unavailable account_id=%s post_id=%s error=%s",
            account.id,
            post_id,
            _safe_error(error),
        )

    metrics.update(_fetch_facebook_post_insight_metrics(account, post_id))
    if "engagement" not in metrics and any(key in metrics for key in ("comments", "shares", "reactions", "clicks")):
        metrics["engagement"] = sum(float(metrics.get(key) or 0) for key in ("comments", "shares", "reactions", "clicks"))
    return metrics


def _fetch_instagram_media_metrics(account: SocialAccount, media_id: str) -> dict[str, float]:
    if not account.access_token:
        return {}
    metric_names = POST_INSIGHT_METRICS.get("instagram", [])
    if not metric_names:
        return {}

    try:
        payload = _graph_get(
            f"{media_id}/insights",
            account.access_token,
            {"metric": ",".join(metric_names)},
        )
        metrics = _media_insight_values_by_name(payload)
        if metrics:
            return metrics
    except Exception as error:
        logger.info(
            "Instagram combined media insights unavailable account_id=%s media_id=%s error=%s",
            account.id,
            media_id,
            _safe_error(error),
        )

    fallback_metrics: dict[str, float] = {}
    for metric_name in metric_names:
        try:
            payload = _graph_get(
                f"{media_id}/insights",
                account.access_token,
                {"metric": metric_name},
            )
        except Exception as error:
            logger.info(
                "Instagram media insight unavailable account_id=%s media_id=%s metric=%s error=%s",
                account.id,
                media_id,
                metric_name,
                _safe_error(error),
            )
            continue
        metric_value = _media_insight_value(payload)
        if metric_value is not None:
            fallback_metrics[metric_name] = metric_value
    return fallback_metrics


def _instagram_media_field_metrics(media: dict[str, Any]) -> dict[str, float]:
    metrics: dict[str, float] = {}
    likes = _metric_number(media.get("like_count")) if isinstance(media, dict) else None
    if likes is not None:
        metrics["likes"] = likes
    comments = _metric_number(media.get("comments_count")) if isinstance(media, dict) else None
    if comments is not None:
        metrics["comments"] = comments
    return metrics


def _refresh_post_insights_for_reference(reference: PlatformPostReference, account: SocialAccount) -> list[PostInsightSnapshot]:
    metric_names = POST_INSIGHT_METRICS.get(reference.platform, [])
    if not metric_names or not account.access_token or not reference.platform_post_id:
        return []

    stored: list[PostInsightSnapshot] = []
    date_value = reference.published_at
    try:
        if reference.platform == "facebook":
            metric_values = _fetch_facebook_post_metrics(account, reference.platform_post_id)
        elif reference.platform == "instagram":
            metric_values = _fetch_instagram_media_metrics(account, reference.platform_post_id)
        else:
            metric_values = {}
    except Exception as error:
        for metric_name in metric_names:
            stored.append(
                _upsert_post_insight(
                    reference,
                    metric_name=metric_name,
                    metric_value=None,
                    period="lifetime",
                    date_value=date_value,
                    status="error" if not _is_permission_error(str(error)) else "unavailable",
                    error_message=_safe_error(error),
                )
            )
        return stored

    for metric_name in metric_names:
        metric_value = metric_values.get(metric_name)
        stored.append(
            _upsert_post_insight(
                reference,
                metric_name=metric_name,
                metric_value=metric_value,
                period="lifetime",
                date_value=date_value,
                status="ok" if metric_value is not None else "unavailable",
                error_message=None,
            )
        )
    return stored


def _refresh_recent_post_insights(account: SocialAccount) -> list[PostInsightSnapshot]:
    references = _ensure_post_references_for_account(account)
    cutoff = utcnow() - timedelta(days=POST_INSIGHTS_LOOKBACK_DAYS)
    stored: list[PostInsightSnapshot] = []
    skipped_remote_ids: set[str] = set()
    for reference in references:
        if reference.published_at and reference.published_at < cutoff:
            continue
        skipped_remote_ids.add(reference.platform_post_id)
        stored.extend(_refresh_post_insights_for_reference(reference, account))
    discovery_limit = 250
    if account.platform == "instagram":
        stored.extend(
            _recent_instagram_media_rows(
                account,
                since=cutoff,
                limit=discovery_limit,
                skipped_media_ids=skipped_remote_ids,
            )
        )
    elif account.platform == "facebook":
        stored.extend(
            _recent_facebook_post_rows(
                account,
                since=cutoff,
                limit=discovery_limit,
                skipped_post_ids=skipped_remote_ids,
            )
        )
    return stored


def refresh_account_insights(
    account: SocialAccount,
    *,
    force: bool = False,
    start: datetime | None = None,
    end: datetime | None = None,
) -> dict[str, Any]:
    logger.info(
        "Starting social insights refresh account_id=%s platform=%s force=%s",
        account.id,
        account.platform,
        force,
    )
    if account.platform not in {"facebook", "instagram"}:
        return {"account_id": account.id, "status": "skipped", "reason": "unsupported platform"}
    if not account.is_active:
        return {"account_id": account.id, "status": "skipped", "reason": "inactive"}

    latest = (
        SocialInsight.query.filter_by(social_account_id=account.id)
        .order_by(SocialInsight.refreshed_at.desc())
        .first()
    )
    if latest and not force:
        age_seconds = (utcnow() - latest.refreshed_at).total_seconds()
        if age_seconds < SOCIAL_INSIGHTS_MIN_REFRESH_SECONDS:
            return {"account_id": account.id, "status": "skipped", "reason": "recently refreshed"}

    run = _new_run_context()
    default_start, default_end = _date_range(POST_INSIGHTS_LOOKBACK_DAYS)
    start = start or default_start
    end = end or default_end
    if end < start:
        return {"account_id": account.id, "status": "failed", "error": "Refresh end date must be after the start date."}
    count = 0
    post_count = 0
    metric_errors = 0
    try:
        _cleanup_removed_insight_metrics()
        _ensure_meta_token(account)
        _cleanup_instagram_legacy_follower_diagnostics(account)
        if not account.page_id_external or not account.access_token:
            message = "Account is missing Meta object ID or access token. Reconnect the account or refresh the global Meta token."
            _store_account_error(account, message, run)
            db.session.commit()
            return {"account_id": account.id, "status": "failed", "error": message}

        for insight in _refresh_field_metrics(account, run):
            if insight.error_message:
                metric_errors += 1
            elif insight.metric_value is not None:
                count += 1

        for metric_name, graph_metrics in _daily_metric_groups_for_platform(account.platform).items():
            try:
                for insight in _refresh_daily_metric_group(account, metric_name, graph_metrics, start, end, run):
                    if insight.error_message:
                        metric_errors += 1
                    elif insight.metric_value is not None:
                        count += 1
            except Exception as metric_error:
                metric_errors += 1
                _store_metric_error(
                    account,
                    metric_name=metric_name,
                    period="day",
                    start_date=start,
                    end_date=end,
                    metadata={
                        "source": "insights",
                        "candidate_metrics": [str(candidate.get("metric") or "") for candidate in graph_metrics],
                    },
                    message=_safe_error(metric_error),
                    run=run,
                )

        for snapshot in _refresh_recent_post_insights(account):
            if snapshot.status == "error":
                metric_errors += 1
            elif snapshot.metric_value is not None:
                post_count += 1

        if count or post_count:
            account.test_status = "success"
            account.test_error = None
        else:
            account.test_status = "failed"
            account.test_error = "No insights could be refreshed. Check Meta permissions, object ID, and account type."
        account.last_tested = utcnow()
        db.session.commit()
        total_count = count + post_count
        status = "refreshed" if total_count and metric_errors == 0 else "partial" if total_count else "failed"
        return {
            "account_id": account.id,
            "status": status,
            "insights": count,
            "post_insights": post_count,
            "metric_errors": metric_errors,
        }
    except Exception as error:
        db.session.rollback()
        _store_account_error(account, _safe_error(error), run)
        db.session.commit()
        return {"account_id": account.id, "status": "failed", "error": _safe_error(error)}


def refresh_all_social_insights(
    *,
    force: bool = False,
    paced: bool = False,
    start: datetime | None = None,
    end: datetime | None = None,
    progress_callback: AnalyticsProgressCallback | None = None,
) -> dict[str, Any]:
    accounts = (
        SocialAccount.query.filter(SocialAccount.platform.in_(["facebook", "instagram"]))
        .order_by(SocialAccount.id.asc())
        .all()
    )
    results: list[dict[str, Any]] = []
    pace_seconds = max(float(SOCIAL_INSIGHTS_ACCOUNT_PACE_SECONDS or 0), 0.0) if paced else 0.0
    logger.info(
        "Starting social insights refresh for %s account(s); force=%s paced=%s.",
        len(accounts),
        force,
        paced,
    )
    for index, account in enumerate(accounts):
        if progress_callback:
            progress_callback(index + 1, len(accounts), account)
        results.append(refresh_account_insights(account, force=force, start=start, end=end))
        if pace_seconds and index < len(accounts) - 1:
            logger.debug(
                "Pacing social insights refresh for %.1f second(s) before the next account.",
                pace_seconds,
            )
            time.sleep(pace_seconds)
    result = {
        "accounts_seen": len(accounts),
        "refreshed": sum(1 for item in results if item.get("status") in {"refreshed", "partial"}),
        "failed": sum(1 for item in results if item.get("status") == "failed"),
        "paced": paced,
        "pace_seconds": pace_seconds,
        "results": results,
    }
    logger.info("Social insights refresh result: %s", result)
    return result


def _account_payload(account: SocialAccount) -> dict[str, Any]:
    insights = _visible_insight_history_for_account(account.id, limit=5000)
    diagnostics = _diagnostic_insights_for_account(account.id)
    latest = (
        SocialInsight.query.filter_by(social_account_id=account.id)
        .order_by(SocialInsight.refreshed_at.desc())
        .first()
    )
    metrics = _metric_summary_from_insights(insights)
    state = "ready" if account.is_active and account.page_id_external and account.access_token else "needs_setup"
    if account.test_error:
        state = "warning"
    return {
        "id": account.id,
        "page_id": account.page_id,
        "page_name": account.page.name if account.page else None,
        "platform": account.platform,
        "account_name": account.account_name,
        "page_id_external": account.page_id_external,
        "is_active": account.is_active,
        "ready": bool(account.is_active and account.page_id_external and account.access_token),
        "last_refreshed_at": latest.refreshed_at.isoformat() if latest else None,
        "last_refresh_run_id": latest.refresh_run_id if latest else None,
        "last_refresh_run_started_at": latest.refresh_run_started_at.isoformat() if latest and latest.refresh_run_started_at else None,
        "last_error": account.test_error,
        "diagnostics": [insight.to_dict() for insight in diagnostics],
        "state": state,
        "followers": metrics.get("followers", 0),
        "views": metrics.get("views", 0),
        "engagement": metrics.get("engagement", 0),
        "reach": metrics.get("reach", 0),
        "media_count": metrics.get("media_count", 0),
        "insight_count": len(insights),
        "insights": [insight.to_dict() for insight in insights],
    }


def _metric_category(metric_name: str | None) -> str:
    normalized = str(metric_name or "").strip().lower()
    if normalized in {"page_media_view", "views"}:
        return "views"
    if normalized in {"page_post_engagements", "engagement", "total_interactions", "accounts_engaged"}:
        return "engagement"
    if normalized in {"followers", "followers_count", "follower_count"}:
        return "followers"
    if normalized in {"reach", "page_impressions_unique", "page_total_media_view_unique"}:
        return "reach"
    if normalized in {"media_count"}:
        return "media_count"
    if normalized in {"visits", "profile_views", "page_views_total"}:
        return "visits"
    if normalized in {"reactions"}:
        return normalized
    return normalized or "metric"


def _metric_summary_from_insights(insights: list[SocialInsight]) -> dict[str, int]:
    totals: dict[str, float] = {}
    latest_lifetime: dict[str, tuple[datetime, float]] = {}
    latest_snapshot_metrics = {"followers", "media_count"}
    for insight in insights:
        if insight.metric_value is None or insight.error_message:
            continue
        category = _metric_category(insight.metric_name)
        value = float(insight.metric_value or 0)
        if insight.period == "lifetime" or (insight.period == "snapshot" and category in latest_snapshot_metrics):
            marker = insight.refreshed_at or insight.end_date or insight.start_date or utcnow()
            current = latest_lifetime.get(category)
            if not current or marker > current[0]:
                latest_lifetime[category] = (marker, value)
            continue
        totals[category] = totals.get(category, 0) + value
    for category, (_, value) in latest_lifetime.items():
        totals[category] = value
    return {key: int(round(value)) for key, value in totals.items()}


def _parse_range_args() -> tuple[datetime | None, datetime | None]:
    range_key = (request.args.get("range") or "60d").strip().lower()
    now = utcnow()
    start: datetime | None = None
    end: datetime | None = None
    if range_key == "7d":
        start = now - timedelta(days=7)
    elif range_key == "60d":
        start = now - timedelta(days=60)
    elif range_key == "30d":
        start = now - timedelta(days=30)
    elif range_key == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == "custom":
        try:
            start = core.parse_iso_datetime(f"{request.args.get('start')}T00:00:00") if request.args.get("start") else None
        except Exception:
            start = None
        try:
            end = core.parse_iso_datetime(f"{request.args.get('end')}T23:59:59") if request.args.get("end") else None
        except Exception:
            end = None
    return start, end


def _date_in_range(value: datetime | None, start: datetime | None, end: datetime | None) -> bool:
    if not value:
        return start is None and end is None
    if start and value < start:
        return False
    if end and value > end:
        return False
    return True


def _filtered_accounts() -> list[SocialAccount]:
    platform = (request.args.get("platform") or "all").strip().lower()
    page_id = request.args.get("page_id", type=int)
    query = SocialAccount.query.filter(SocialAccount.platform.in_(["facebook", "instagram"]))
    if platform in {"facebook", "instagram"}:
        query = query.filter(SocialAccount.platform == platform)
    if page_id:
        query = query.filter(SocialAccount.page_id == page_id)
    return query.order_by(SocialAccount.platform.asc(), SocialAccount.account_name.asc()).all()


def _filtered_visible_insights(account: SocialAccount) -> list[SocialInsight]:
    start, end = _parse_range_args()
    metric = (request.args.get("metric") or "all").strip().lower()
    rows = _latest_visible_insights_for_account(account.id, limit=1000)
    filtered: list[SocialInsight] = []
    for insight in rows:
        if metric != "all" and _metric_category(insight.metric_name) != metric and insight.metric_name != metric:
            continue
        if not _date_in_range(_insight_reporting_date(insight) or insight.refreshed_at, start, end):
            continue
        filtered.append(insight)
    return filtered


def _trend_rows(accounts: list[SocialAccount]) -> list[dict[str, Any]]:
    buckets: dict[str, dict[str, float]] = {}
    for account in accounts:
        for insight in _filtered_visible_insights(account):
            if insight.metric_value is None:
                continue
            date_value = _insight_reporting_date(insight) or insight.refreshed_at
            if not date_value:
                continue
            key = date_value.strftime("%Y-%m-%d")
            bucket = buckets.setdefault(
                key,
                {"date": key, "views": 0, "engagement": 0, "followers": 0, "reach": 0, "visits": 0, "media_count": 0},
            )
            category = _metric_category(insight.metric_name)
            if category in bucket:
                bucket[category] = float(bucket.get(category, 0)) + float(insight.metric_value or 0)
    return [
        {
            key: int(round(value)) if isinstance(value, float) else value
            for key, value in bucket.items()
        }
        for _, bucket in sorted(buckets.items())
    ]


def _scheduler_refresh_info() -> dict[str, Any]:
    job = core.scheduler.get_job("refresh_social_insights") if core.scheduler.running else None
    latest = db.session.query(func.max(SocialInsight.refreshed_at)).scalar()
    return {
        "status": "idle",
        "last_refreshed_at": latest.isoformat() if latest else None,
        "next_scheduled_refresh_at": job.next_run_time.isoformat() if job and job.next_run_time else None,
        "refresh_interval_seconds": SOCIAL_INSIGHTS_REFRESH_INTERVAL_SECONDS,
    }


def _report_text_key(value: Any) -> str:
    text = str(value or "").strip().lower().replace("\xa0", " ")
    return "".join(character for character in text if character.isalnum())


def _report_template_path() -> Path:
    configured_path = os.environ.get(REPORT_TEMPLATE_PATH_ENV, "").strip()
    template_path = Path(configured_path or DEFAULT_REPORT_TEMPLATE_PATH).expanduser()
    if not template_path.exists():
        raise RuntimeError(
            f"Report template not found at {template_path}. "
            f"Set {REPORT_TEMPLATE_PATH_ENV} to the workbook template path."
        )
    return template_path


def _google_report_credentials_path() -> Path:
    configured_path = (
        os.environ.get(GOOGLE_REPORT_CREDENTIALS_PATH_ENV, "").strip()
        or os.environ.get(GOOGLE_APPLICATION_CREDENTIALS_ENV, "").strip()
    )
    credentials_path = Path(configured_path or DEFAULT_GOOGLE_REPORT_CREDENTIALS_PATH).expanduser()
    if not credentials_path.exists():
        raise RuntimeError(
            f"Google service account credentials not found at {credentials_path}. "
            f"Set {GOOGLE_REPORT_CREDENTIALS_PATH_ENV} or {GOOGLE_APPLICATION_CREDENTIALS_ENV}."
        )
    return credentials_path


def _google_report_spreadsheet_id() -> str:
    spreadsheet_id = os.environ.get(GOOGLE_REPORT_SPREADSHEET_ID_ENV, "").strip() or DEFAULT_GOOGLE_REPORT_SPREADSHEET_ID
    if not spreadsheet_id:
        raise RuntimeError(f"Set {GOOGLE_REPORT_SPREADSHEET_ID_ENV} to the Marketing Report spreadsheet ID.")
    return spreadsheet_id


def _google_campaign_spreadsheet_id() -> str:
    spreadsheet_id = (
        os.environ.get(GOOGLE_CAMPAIGN_SPREADSHEET_ID_ENV, "").strip()
        or DEFAULT_GOOGLE_CAMPAIGN_SPREADSHEET_ID
    )
    if not spreadsheet_id:
        raise RuntimeError(f"Set {GOOGLE_CAMPAIGN_SPREADSHEET_ID_ENV} to the Campaign spreadsheet ID.")
    return spreadsheet_id


def _google_sheets_service() -> Any:
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError as error:
        raise RuntimeError("google-auth and google-api-python-client are required for Google Sheets report sync.") from error

    credentials = Credentials.from_service_account_file(
        str(_google_report_credentials_path()),
        scopes=list(GOOGLE_SHEETS_SCOPES),
    )
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def _report_year_from_template(template_path: Path) -> int:
    match = re.search(r"(20\d{2})", template_path.name)
    return int(match.group(1)) if match else utcnow().year


def _report_month_window(year: int, month: int) -> tuple[datetime, datetime]:
    start = datetime(year, month, 1)
    if month == 12:
        return start, datetime(year + 1, 1, 1)
    return start, datetime(year, month + 1, 1)


def _report_month_columns(worksheet: Any) -> dict[int, int]:
    columns: dict[int, int] = {}
    for cell in worksheet[1]:
        month = REPORT_MONTH_KEYS.get(_report_text_key(cell.value))
        if month and month not in columns:
            columns[month] = cell.column
    return columns


def _report_social_row(worksheet: Any) -> int | None:
    for row_number in range(1, worksheet.max_row + 1):
        if _report_text_key(worksheet.cell(row_number, 1).value) == "socialmedia":
            return row_number
    return None


def _report_platform_rows(worksheet: Any, social_row: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row_number in range(social_row + 1, worksheet.max_row + 1):
        platform = REPORT_PLATFORM_KEYS.get(_report_text_key(worksheet.cell(row_number, 1).value))
        if platform and platform not in rows:
            rows[platform] = row_number
    return rows


def _report_metric_rows(worksheet: Any, platform_row: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row_number in range(platform_row + 1, worksheet.max_row + 1):
        key = _report_text_key(worksheet.cell(row_number, 1).value)
        if key in REPORT_SECTION_BOUNDARY_KEYS:
            break
        metric = REPORT_METRIC_KEYS.get(key)
        if metric and metric not in rows:
            rows[metric] = row_number
    return rows


def _report_month_columns_from_row(values: list[Any]) -> dict[int, int]:
    columns: dict[int, int] = {}
    for index, value in enumerate(values, start=1):
        month = REPORT_MONTH_KEYS.get(_report_text_key(value))
        if month and month not in columns:
            columns[month] = index
    return columns


def _report_social_row_from_column(values: list[Any]) -> int | None:
    for index, value in enumerate(values, start=1):
        if _report_text_key(value) == "socialmedia":
            return index
    return None


def _report_platform_rows_from_column(values: list[Any], social_row: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for index in range(social_row + 1, len(values) + 1):
        platform = REPORT_PLATFORM_KEYS.get(_report_text_key(values[index - 1]))
        if platform and platform not in rows:
            rows[platform] = index
    return rows


def _report_metric_rows_from_column(values: list[Any], platform_row: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for index in range(platform_row + 1, len(values) + 1):
        key = _report_text_key(values[index - 1])
        if key in REPORT_SECTION_BOUNDARY_KEYS:
            break
        metric = REPORT_METRIC_KEYS.get(key)
        if metric and metric not in rows:
            rows[metric] = index
    return rows


def _report_is_image_posts_key(key: str) -> bool:
    return key in {"imageposts", "imagevideosposts", "imageandvideosposts"} or ("image" in key and "post" in key)


def _report_post_content_rows_from_column(values: list[Any]) -> dict[str, int] | None:
    post_content_row = None
    for index, value in enumerate(values, start=1):
        if _report_text_key(value) == "postcontent":
            post_content_row = index
            break
    if not post_content_row:
        return None

    design_row = None
    image_row = None
    for index in range(post_content_row + 1, len(values) + 1):
        key = _report_text_key(values[index - 1])
        if not design_row and key == "designposts":
            design_row = index
            continue
        if design_row and _report_is_image_posts_key(key):
            image_row = index
            break
    if not design_row or not image_row:
        return None
    return {
        "post_content": post_content_row,
        "design_label": design_row,
        "image_label": image_row,
    }


def _normalized_match_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _match_tokens(value: Any) -> set[str]:
    return {token for token in _normalized_match_text(value).split() if len(token) > 2}


def _campaign_header_map(header_values: list[Any]) -> dict[str, int]:
    return {
        _report_text_key(value): index
        for index, value in enumerate(header_values)
        if str(value or "").strip()
    }


def _campaign_cell(row: list[Any], headers: dict[str, int], key: str, fallback_index: int | None = None) -> str:
    index = headers.get(key)
    if index is None:
        index = fallback_index
    if index is None or index >= len(row):
        return ""
    return str(row[index] or "").strip()


def _campaign_month_marker(row: list[Any]) -> int | None:
    populated = [str(value).strip() for value in row if str(value or "").strip()]
    if len(populated) != 1:
        return None
    direct = REPORT_MONTH_KEYS.get(_report_text_key(populated[0]))
    if direct:
        return direct
    prefix = re.match(r"^\s*([A-Za-z]+)\b", populated[0])
    return REPORT_MONTH_KEYS.get(_report_text_key(prefix.group(1))) if prefix else None


def _campaign_date_value(value: Any, target_year: int) -> datetime | None:
    match = re.search(r"\b(\d{1,2})\s+([A-Za-z]+)\b", str(value or "").replace(",", " "))
    if not match:
        return None
    month_text = _report_text_key(match.group(2))
    month = REPORT_MONTH_KEYS.get(month_text) or REPORT_MONTH_KEYS.get(month_text[:3])
    if not month:
        return None
    try:
        return datetime(target_year, month, int(match.group(1)))
    except ValueError:
        return None


def _campaign_post_category(row: list[Any], headers: dict[str, int]) -> str:
    format_text = _normalized_match_text(_campaign_cell(row, headers, "format", 7))
    creative_text = _normalized_match_text(_campaign_cell(row, headers, "creative", 11))
    notes_text = " ".join(
        [
            _normalized_match_text(_campaign_cell(row, headers, "mssnotes")),
            _normalized_match_text(_campaign_cell(row, headers, "deadlines")),
        ]
    )
    if re.search(r"\b(vid|video|reel)\b", format_text):
        return "image"
    if re.search(r"\b(whatsapp|photo|photos|image|images|vid|video|reel)\b", creative_text):
        return "image"
    if re.search(r"\b(whatsapp|photo|photos|image|images|vid|video|reel)\b", notes_text):
        return "image"
    return "design"


def _campaign_posts_for_month(
    values: list[list[Any]],
    target_year: int,
    target_month: int,
) -> list[dict[str, Any]]:
    if not values:
        return []

    headers = _campaign_header_map(values[0])
    start_index = None
    for index, row in enumerate(values, start=1):
        if _campaign_month_marker(row) == target_month:
            start_index = index + 1
            break
    if not start_index:
        return []

    rows: list[dict[str, Any]] = []
    blank_streak = 0
    for row_number in range(start_index, len(values) + 1):
        row = values[row_number - 1]
        if _campaign_month_marker(row):
            break
        if not any(str(value or "").strip() for value in row):
            blank_streak += 1
            if blank_streak >= 3 and rows:
                break
            continue
        blank_streak = 0

        theme = _campaign_cell(row, headers, "theme", 4)
        if not theme:
            continue
        post_copy = _campaign_cell(row, headers, "postcopy", 5)
        date_value = _campaign_date_value(_campaign_cell(row, headers, "date", 2), target_year)
        rows.append(
            {
                "row": row_number,
                "theme": theme,
                "post_copy": post_copy,
                "post_copy_norm": _normalized_match_text(post_copy),
                "post_copy_tokens": _match_tokens(post_copy),
                "date": date_value,
                "category": _campaign_post_category(row, headers),
            }
        )
    return rows


def _post_reference_latest_metric(reference: PlatformPostReference, metric_name: str) -> int:
    snapshot = (
        PostInsightSnapshot.query.filter_by(
            internal_post_id=reference.internal_post_id,
            social_account_id=reference.social_account_id,
            platform_post_id=reference.platform_post_id,
            metric_name=metric_name,
        )
        .filter(PostInsightSnapshot.status.in_(["ok", "stale"]))
        .order_by(PostInsightSnapshot.fetched_at.desc().nullslast(), PostInsightSnapshot.id.desc())
        .first()
    )
    if not snapshot or snapshot.metric_value is None:
        return 0
    return int(round(float(snapshot.metric_value or 0)))


def _report_post_groups_for_page(
    page: Page,
    target_start: datetime,
    target_end: datetime,
) -> list[dict[str, Any]]:
    accounts = (
        SocialAccount.query.filter_by(page_id=page.id)
        .filter(SocialAccount.platform.in_(["facebook", "instagram"]))
        .order_by(SocialAccount.id.asc())
        .all()
    )
    account_ids = [account.id for account in accounts]
    if not account_ids:
        return []

    reference_start = target_start - timedelta(days=CAMPAIGN_POST_MATCH_PAD_DAYS)
    reference_end = target_end + timedelta(days=CAMPAIGN_POST_MATCH_PAD_DAYS)
    references = (
        PlatformPostReference.query.filter(PlatformPostReference.social_account_id.in_(account_ids))
        .filter(PlatformPostReference.published_at >= reference_start)
        .filter(PlatformPostReference.published_at <= reference_end)
        .order_by(PlatformPostReference.published_at.asc().nullslast(), PlatformPostReference.id.asc())
        .all()
    )

    grouped: dict[int, dict[str, Any]] = {}
    seen_metric_refs: set[tuple[int, int, str, str]] = set()
    for reference in references:
        group = grouped.setdefault(
            reference.internal_post_id,
            {
                "internal_post_id": reference.internal_post_id,
                "views": 0,
                "dates": [],
                "captions": [],
                "platforms": set(),
            },
        )
        metric_key = (
            reference.internal_post_id,
            reference.social_account_id,
            reference.platform,
            reference.platform_post_id,
        )
        if metric_key not in seen_metric_refs:
            seen_metric_refs.add(metric_key)
            group["views"] += _post_reference_latest_metric(reference, "views")
        if reference.published_at:
            group["dates"].append(reference.published_at)
        group["platforms"].add(reference.platform)
        caption = reference.caption_preview or (reference.post.content if reference.post else "") or ""
        if caption:
            group["captions"].append(caption)

    groups: list[dict[str, Any]] = []
    for group in grouped.values():
        caption = max(group["captions"], key=len, default="")
        published_at = min(group["dates"]) if group["dates"] else None
        groups.append(
            {
                "internal_post_id": group["internal_post_id"],
                "views": int(group["views"]),
                "caption": caption,
                "caption_norm": _normalized_match_text(caption),
                "caption_tokens": _match_tokens(caption),
                "published_at": published_at,
                "platforms": sorted(group["platforms"]),
            }
        )
    return groups


def _campaign_post_match_score(campaign_row: dict[str, Any], post_group: dict[str, Any]) -> float:
    post_copy_norm = campaign_row.get("post_copy_norm") or ""
    caption_norm = post_group.get("caption_norm") or ""
    if not post_copy_norm or not caption_norm:
        base_score = 0.0
    elif post_copy_norm in caption_norm or caption_norm in post_copy_norm:
        base_score = 1.0
    else:
        base_score = difflib.SequenceMatcher(None, post_copy_norm[:650], caption_norm[:650]).ratio()
        campaign_tokens = campaign_row.get("post_copy_tokens") or set()
        caption_tokens = post_group.get("caption_tokens") or set()
        denominator = max(min(len(campaign_tokens), len(caption_tokens)), 1)
        base_score = max(base_score, len(campaign_tokens & caption_tokens) / denominator)

    date_value = campaign_row.get("date")
    published_at = post_group.get("published_at")
    if not date_value or not published_at:
        return base_score

    day_delta = abs((published_at.date() - date_value.date()).days)
    if day_delta == 0:
        return base_score + 0.25
    if day_delta == 1:
        return base_score + 0.18
    if day_delta == 2:
        return base_score + 0.08
    if day_delta > 7:
        return base_score - 0.15
    return base_score


def _match_campaign_posts_to_groups(
    campaign_rows: list[dict[str, Any]],
    post_groups: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    matched: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    used_post_ids: set[int] = set()
    for campaign_row in campaign_rows:
        if not campaign_row.get("post_copy_norm"):
            unmatched.append({**campaign_row, "reason": "missing_post_copy", "score": 0})
            continue
        best_score = 0.0
        best_group = None
        for post_group in post_groups:
            if post_group["internal_post_id"] in used_post_ids:
                continue
            score = _campaign_post_match_score(campaign_row, post_group)
            if score > best_score:
                best_score = score
                best_group = post_group
        if best_group and best_score >= CAMPAIGN_POST_MATCH_MIN_SCORE:
            used_post_ids.add(best_group["internal_post_id"])
            matched.append(
                {
                    **campaign_row,
                    "views": int(best_group["views"]),
                    "score": round(best_score, 3),
                    "internal_post_id": best_group["internal_post_id"],
                    "platforms": best_group["platforms"],
                    "published_at": best_group["published_at"],
                }
            )
        else:
            unmatched.append(
                {
                    **campaign_row,
                    "reason": "no_matching_post_metric",
                    "score": round(best_score, 3),
                }
            )
    return matched, unmatched


def _sheet_a1_quote(title: str) -> str:
    escaped = title.replace("'", "''")
    return f"'{escaped}'"


def _sheet_column_name(index: int) -> str:
    if index < 1:
        raise ValueError("Column index must be positive.")
    result = ""
    cursor = index
    while cursor:
        cursor, remainder = divmod(cursor - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _report_pages_by_key() -> dict[str, Page]:
    return {_report_text_key(page.name): page for page in Page.query.order_by(Page.name.asc()).all()}


def _report_page_for_sheet(sheet_name: str, pages_by_key: dict[str, Page]) -> Page | None:
    sheet_key = _report_text_key(sheet_name)
    if sheet_key in REPORT_IGNORED_SHEET_KEYS:
        return None
    alias = REPORT_SHEET_PAGE_ALIASES.get(sheet_key)
    return pages_by_key.get(_report_text_key(alias)) if alias else pages_by_key.get(sheet_key)


def _report_campaign_sheet_for_report_sheet(sheet_name: str, campaign_titles_by_key: dict[str, str]) -> str | None:
    sheet_key = _report_text_key(sheet_name)
    if sheet_key in REPORT_IGNORED_SHEET_KEYS:
        return None
    alias = REPORT_CAMPAIGN_SHEET_ALIASES.get(sheet_key)
    return campaign_titles_by_key.get(_report_text_key(alias)) if alias else campaign_titles_by_key.get(sheet_key)


def _report_account_for_platform(page: Page, platform: str) -> SocialAccount | None:
    return (
        SocialAccount.query.filter_by(page_id=page.id, platform=platform)
        .order_by(SocialAccount.id.asc())
        .first()
    )


def _report_metric_groups_for_platform(platform: str) -> dict[str, list[InsightMetricCandidate]]:
    if platform == "facebook":
        return REPORT_FACEBOOK_METRIC_GROUPS
    if platform == "instagram":
        return REPORT_INSTAGRAM_METRIC_GROUPS
    return {}


def _report_previous_completed_month(reference: datetime | None = None) -> tuple[int, int, datetime, datetime]:
    reference = reference or utcnow()
    current_month_start = reference.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    target_end = current_month_start - timedelta(seconds=1)
    target_start = target_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return target_start.year, target_start.month, target_start, target_end


def _report_snapshot_date(insight: SocialInsight) -> datetime | None:
    return _insight_reporting_date(insight) or insight.start_date or insight.end_date


def _report_select_interval_partition(
    candidates: list[tuple[datetime, datetime, float, datetime]],
    target_start: datetime,
    target_end: datetime,
) -> float | None:
    latest_by_range: dict[tuple[datetime, datetime], tuple[datetime, datetime, float, datetime]] = {}
    for candidate in candidates:
        key = (candidate[0], candidate[1])
        current = latest_by_range.get(key)
        if not current or candidate[3] > current[3]:
            latest_by_range[key] = candidate

    exact = [
        candidate
        for candidate in latest_by_range.values()
        if candidate[0] == target_start and candidate[1] == target_end
    ]
    if exact:
        return max(exact, key=lambda candidate: candidate[3])[2]

    cursor = target_start
    selected: list[tuple[datetime, datetime, float, datetime]] = []
    while cursor <= target_end:
        choices = [
            candidate
            for candidate in latest_by_range.values()
            if candidate[0] == cursor and cursor <= candidate[1] <= target_end
        ]
        if not choices:
            return None
        choice = max(choices, key=lambda candidate: (candidate[1], candidate[3]))
        selected.append(choice)
        cursor = choice[1] + timedelta(days=1)

    return sum(candidate[2] for candidate in selected)


def _report_activity_value(insights: list[SocialInsight], metric: str, target_start: datetime, target_end: datetime) -> int | None:
    target_start_day = _start_of_day(target_start)
    target_end_day = _start_of_day(target_end)
    interval_candidates: list[tuple[datetime, datetime, float, datetime]] = []
    daily_values: dict[datetime, tuple[datetime, float]] = {}
    single_period_values: list[tuple[datetime, float]] = []

    for insight in insights:
        if _metric_category(insight.metric_name) != metric or insight.metric_value is None:
            continue
        value = float(insight.metric_value or 0)
        marker = insight.refreshed_at or target_end
        if insight.start_date and insight.end_date:
            start_day = _start_of_day(insight.start_date)
            end_day = _start_of_day(insight.end_date)
            if target_start_day <= start_day <= end_day <= target_end_day:
                interval_candidates.append((start_day, end_day, value, marker))
            continue

        date_value = _insight_reporting_date(insight) or insight.start_date or insight.end_date
        if not date_value:
            continue
        day = _start_of_day(date_value)
        if not target_start_day <= day <= target_end_day:
            continue
        if metric in REPORT_NON_ADDITIVE_ACTIVITY_METRICS and str(insight.period or "").lower() != "day":
            single_period_values.append((day, value))
            continue
        current = daily_values.get(day)
        if not current or marker > current[0]:
            daily_values[day] = (marker, value)

    interval_total = _report_select_interval_partition(interval_candidates, target_start_day, target_end_day)
    if interval_total is not None:
        return int(round(interval_total))
    if metric in REPORT_NON_ADDITIVE_ACTIVITY_METRICS:
        if single_period_values:
            return int(round(max(single_period_values, key=lambda item: item[0])[1]))
        if daily_values:
            return int(round(sum(value for _, value in daily_values.values())))
        return None
    if daily_values:
        return int(round(sum(value for _, value in daily_values.values())))
    return None


def _report_snapshot_value(insights: list[SocialInsight], metric: str, target_start: datetime, target_end: datetime) -> int | None:
    dated_candidates: list[tuple[datetime, float]] = []
    fallback_candidates: list[tuple[datetime, float]] = []
    for insight in insights:
        if _metric_category(insight.metric_name) != metric or insight.metric_value is None:
            continue
        value = float(insight.metric_value or 0)
        snapshot_date = _report_snapshot_date(insight)
        if snapshot_date:
            dated_candidates.append((snapshot_date, value))
        elif insight.refreshed_at:
            fallback_candidates.append((insight.refreshed_at, value))

    if not dated_candidates and not fallback_candidates:
        return None

    grace_end = target_end + timedelta(days=7)
    in_month = [item for item in dated_candidates if target_start <= item[0] <= target_end]
    before_end = [item for item in dated_candidates if item[0] <= target_end]
    shortly_after_end = [item for item in dated_candidates if target_end < item[0] <= grace_end]
    selected = (
        max(in_month, key=lambda item: item[0])
        if in_month
        else max(before_end, key=lambda item: item[0])
        if before_end
        else min(shortly_after_end, key=lambda item: item[0])
        if shortly_after_end
        else None
    )
    if selected:
        return int(round(selected[1]))

    fallback_in_window = [item for item in fallback_candidates if item[0] <= grace_end]
    if fallback_in_window:
        return int(round(max(fallback_in_window, key=lambda item: item[0])[1]))
    return None


def _report_account_month_metric_values(
    account: SocialAccount,
    target_start: datetime,
    target_end: datetime,
) -> dict[str, int]:
    report_metrics = set(REPORT_METRIC_KEYS.values())
    insights = (
        SocialInsight.query.filter_by(social_account_id=account.id)
        .filter(*_visible_insight_filter())
        .order_by(SocialInsight.start_date.asc().nullsfirst(), SocialInsight.end_date.asc().nullsfirst())
        .all()
    )
    values: dict[str, int] = {}
    for metric in report_metrics:
        value = (
            _report_snapshot_value(insights, metric, target_start, target_end)
            if metric in REPORT_SNAPSHOT_METRICS
            else _report_activity_value(insights, metric, target_start, target_end)
        )
        if value is not None:
            values[metric] = value
    return values


def _report_post_content_value_ranges(
    sheet_title: str,
    target_column: int,
    section_rows: dict[str, int],
    matched_rows: list[dict[str, Any]],
    inserted_design_rows: int,
) -> list[dict[str, Any]]:
    design_entries = [row for row in matched_rows if row.get("category") == "design"]
    image_entries = [row for row in matched_rows if row.get("category") != "design"]

    design_start = section_rows["design_label"] + 1
    image_label = section_rows["image_label"] + inserted_design_rows
    image_start = image_label + 1
    design_clear_rows = max(image_label - design_start, len(design_entries), 1)
    image_clear_rows = max(REPORT_POST_CONTENT_CLEAR_ROWS, len(image_entries), 1)

    theme_column = _sheet_column_name(target_column)
    views_column = _sheet_column_name(target_column + 1)
    quoted_title = _sheet_a1_quote(sheet_title)

    design_values = [
        [entry.get("theme") or "", int(entry.get("views") or 0)]
        for entry in design_entries
    ]
    design_values.extend([["", ""] for _ in range(design_clear_rows - len(design_values))])

    image_values = [
        [entry.get("theme") or "", int(entry.get("views") or 0)]
        for entry in image_entries
    ]
    image_values.extend([["", ""] for _ in range(image_clear_rows - len(image_values))])

    return [
        {
            "range": f"{quoted_title}!{theme_column}{design_start}:{views_column}{design_start + design_clear_rows - 1}",
            "values": design_values,
        },
        {
            "range": f"{quoted_title}!{theme_column}{image_start}:{views_column}{image_start + image_clear_rows - 1}",
            "values": image_values,
        },
    ]


def _insert_report_rows(
    service: Any,
    spreadsheet_id: str,
    sheet_id: int,
    before_row: int,
    row_count: int,
) -> None:
    if row_count <= 0:
        return
    start_index = max(before_row - 1, 0)
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "insertDimension": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "ROWS",
                            "startIndex": start_index,
                            "endIndex": start_index + row_count,
                        },
                        "inheritFromBefore": start_index > 0,
                    }
                }
            ]
        },
    ).execute()


def _report_post_row_summary(row: dict[str, Any]) -> dict[str, Any]:
    date_value = row.get("date")
    published_at = row.get("published_at")
    return {
        "campaign_row": row.get("row"),
        "theme": row.get("theme"),
        "category": row.get("category"),
        "views": row.get("views"),
        "score": row.get("score"),
        "reason": row.get("reason"),
        "internal_post_id": row.get("internal_post_id"),
        "date": date_value.date().isoformat() if isinstance(date_value, datetime) else None,
        "published_at": published_at.isoformat() if isinstance(published_at, datetime) else None,
    }


def _refresh_report_month_insights(target_start: datetime, target_end: datetime) -> dict[str, Any]:
    accounts = (
        SocialAccount.query.filter(SocialAccount.platform.in_(["facebook", "instagram"]))
        .order_by(SocialAccount.id.asc())
        .all()
    )
    result = {"accounts_seen": len(accounts), "refreshed": 0, "failed": 0, "metric_errors": 0}
    _cleanup_removed_insight_metrics()
    for account in accounts:
        run = _new_run_context()
        try:
            _ensure_meta_token(account)
            _cleanup_instagram_legacy_follower_diagnostics(account)
            if not account.is_active or not account.page_id_external or not account.access_token:
                continue

            count = 0
            metric_errors = 0
            for insight in _refresh_field_metrics(account, run):
                if insight.error_message:
                    metric_errors += 1
                elif insight.metric_value is not None:
                    count += 1

            for metric_name, graph_metrics in _report_metric_groups_for_platform(account.platform).items():
                for insight in _refresh_daily_metric_group(account, metric_name, graph_metrics, target_start, target_end, run):
                    if insight.error_message:
                        metric_errors += 1
                    elif insight.metric_value is not None:
                        count += 1

            account.test_status = "success" if count else account.test_status
            account.test_error = None if count else account.test_error
            account.last_tested = utcnow()
            db.session.commit()
            result["refreshed"] += 1
            result["metric_errors"] += metric_errors
        except Exception as error:
            db.session.rollback()
            result["failed"] += 1
            logger.warning(
                "Report month insight refresh failed account_id=%s platform=%s: %s",
                account.id,
                account.platform,
                _safe_error(error),
            )
            continue
    return result


def update_marketing_report_google_sheet(
    target_year: int | None = None,
    target_month: int | None = None,
    target_start: datetime | None = None,
    target_end: datetime | None = None,
    *,
    dry_run: bool = False,
) -> dict[str, Any]:
    if not target_year or not target_month or not target_start or not target_end:
        target_year, target_month, target_start, target_end = _report_previous_completed_month()

    service = _google_sheets_service()
    spreadsheet_id = _google_report_spreadsheet_id()
    spreadsheet = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="properties(title),sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)))",
    ).execute()
    sheet_properties_by_title: dict[str, dict[str, Any]] = {}
    sheet_titles: list[str] = []
    for sheet in spreadsheet.get("sheets", []):
        properties = sheet.get("properties", {})
        title = str(properties.get("title") or "")
        if not title:
            continue
        sheet_titles.append(title)
        sheet_properties_by_title[title] = properties
    if not sheet_titles:
        raise RuntimeError("The Google report spreadsheet has no visible sheets.")

    campaign_spreadsheet_id = _google_campaign_spreadsheet_id()
    campaign_spreadsheet = service.spreadsheets().get(
        spreadsheetId=campaign_spreadsheet_id,
        fields="properties(title),sheets(properties(title))",
    ).execute()
    campaign_titles = [
        str(sheet.get("properties", {}).get("title") or "")
        for sheet in campaign_spreadsheet.get("sheets", [])
        if sheet.get("properties", {}).get("title")
    ]
    campaign_titles_by_key = {_report_text_key(title): title for title in campaign_titles}

    ranges: list[str] = []
    for title in sheet_titles:
        quoted_title = _sheet_a1_quote(title)
        ranges.append(f"{quoted_title}!1:1")
        ranges.append(f"{quoted_title}!A:A")
    value_ranges = service.spreadsheets().values().batchGet(
        spreadsheetId=spreadsheet_id,
        ranges=ranges,
        valueRenderOption="FORMATTED_VALUE",
    ).execute().get("valueRanges", [])

    sheet_values: dict[str, dict[str, list[Any]]] = {}
    for index, title in enumerate(sheet_titles):
        header_values = value_ranges[index * 2].get("values", []) if index * 2 < len(value_ranges) else []
        column_values = value_ranges[index * 2 + 1].get("values", []) if index * 2 + 1 < len(value_ranges) else []
        sheet_values[title] = {
            "header": header_values[0] if header_values else [],
            "column_a": [row[0] if row else "" for row in column_values],
        }

    pages_by_key = _report_pages_by_key()
    campaign_needed_titles: list[str] = []
    for sheet_title in sheet_titles:
        if not _report_page_for_sheet(sheet_title, pages_by_key):
            continue
        campaign_title = _report_campaign_sheet_for_report_sheet(sheet_title, campaign_titles_by_key)
        if campaign_title and campaign_title not in campaign_needed_titles:
            campaign_needed_titles.append(campaign_title)

    campaign_values_by_title: dict[str, list[list[Any]]] = {}
    if campaign_needed_titles:
        campaign_ranges = [f"{_sheet_a1_quote(title)}!A1:AF1200" for title in campaign_needed_titles]
        campaign_value_ranges = service.spreadsheets().values().batchGet(
            spreadsheetId=campaign_spreadsheet_id,
            ranges=campaign_ranges,
            valueRenderOption="FORMATTED_VALUE",
        ).execute().get("valueRanges", [])
        for index, title in enumerate(campaign_needed_titles):
            values = campaign_value_ranges[index].get("values", []) if index < len(campaign_value_ranges) else []
            campaign_values_by_title[title] = values

    account_metric_cache: dict[int, dict[str, int]] = {}
    skipped_sheets: list[str] = []
    update_data: list[dict[str, Any]] = []
    post_content_summary: dict[str, Any] = {
        "campaign_spreadsheet_id": campaign_spreadsheet_id,
        "campaign_spreadsheet_title": campaign_spreadsheet.get("properties", {}).get("title"),
        "campaign_rows": 0,
        "matched_rows": 0,
        "unmatched_rows": 0,
        "written_rows": 0,
        "planned_insert_rows": 0,
        "inserted_rows": 0,
        "sheets": [],
    }

    for sheet_title in sheet_titles:
        page = _report_page_for_sheet(sheet_title, pages_by_key)
        if not page:
            skipped_sheets.append(sheet_title)
            continue

        values = sheet_values.get(sheet_title, {})
        month_columns = _report_month_columns_from_row(values.get("header", []))
        social_row = _report_social_row_from_column(values.get("column_a", []))
        if not month_columns or not social_row:
            skipped_sheets.append(sheet_title)
            continue

        target_column = month_columns.get(target_month)
        if not target_column:
            skipped_sheets.append(sheet_title)
            continue

        platform_rows = _report_platform_rows_from_column(values.get("column_a", []), social_row)
        for platform, platform_row in platform_rows.items():
            account = _report_account_for_platform(page, platform)
            if not account:
                continue
            if account.id not in account_metric_cache:
                account_metric_cache[account.id] = _report_account_month_metric_values(account, target_start, target_end)
            metric_values = account_metric_cache[account.id]
            for metric, row_number in _report_metric_rows_from_column(values.get("column_a", []), platform_row).items():
                if metric not in metric_values:
                    continue
                update_data.append({
                    "range": f"{_sheet_a1_quote(sheet_title)}!{_sheet_column_name(target_column)}{row_number}",
                    "values": [[metric_values[metric]]],
                })

        post_section_rows = _report_post_content_rows_from_column(values.get("column_a", []))
        campaign_title = _report_campaign_sheet_for_report_sheet(sheet_title, campaign_titles_by_key)
        if not post_section_rows or not campaign_title:
            continue

        campaign_rows = _campaign_posts_for_month(
            campaign_values_by_title.get(campaign_title, []),
            target_year,
            target_month,
        )
        post_groups = _report_post_groups_for_page(page, target_start, target_end)
        matched_rows, unmatched_rows = _match_campaign_posts_to_groups(campaign_rows, post_groups)
        design_count = sum(1 for row in matched_rows if row.get("category") == "design")
        image_count = len(matched_rows) - design_count
        design_slots = max(post_section_rows["image_label"] - post_section_rows["design_label"] - 1, 0)
        planned_insert_rows = max(design_count - design_slots, 0)
        inserted_rows = 0
        simulated_insert_rows = planned_insert_rows if dry_run else 0
        if planned_insert_rows and not dry_run:
            sheet_id = sheet_properties_by_title.get(sheet_title, {}).get("sheetId")
            if sheet_id is None:
                raise RuntimeError(f"Could not find the Google sheet ID for {sheet_title}.")
            _insert_report_rows(
                service,
                spreadsheet_id,
                int(sheet_id),
                post_section_rows["image_label"],
                planned_insert_rows,
            )
            inserted_rows = planned_insert_rows
            simulated_insert_rows = inserted_rows

        update_data.extend(
            _report_post_content_value_ranges(
                sheet_title,
                target_column,
                post_section_rows,
                matched_rows,
                simulated_insert_rows,
            )
        )

        post_content_summary["campaign_rows"] += len(campaign_rows)
        post_content_summary["matched_rows"] += len(matched_rows)
        post_content_summary["unmatched_rows"] += len(unmatched_rows)
        post_content_summary["written_rows"] += len(matched_rows)
        post_content_summary["planned_insert_rows"] += planned_insert_rows
        post_content_summary["inserted_rows"] += inserted_rows
        post_content_summary["sheets"].append(
            {
                "sheet": sheet_title,
                "campaign_sheet": campaign_title,
                "campaign_rows": len(campaign_rows),
                "post_reference_groups": len(post_groups),
                "matched_rows": len(matched_rows),
                "unmatched_rows": len(unmatched_rows),
                "design_rows": design_count,
                "image_rows": image_count,
                "design_slots": design_slots,
                "planned_insert_rows": planned_insert_rows,
                "inserted_rows": inserted_rows,
                "matched": [_report_post_row_summary(row) for row in matched_rows],
                "unmatched": [_report_post_row_summary(row) for row in unmatched_rows],
            }
        )

    updated_cells = 0
    if update_data and not dry_run:
        update_result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "valueInputOption": "USER_ENTERED",
                "data": update_data,
            },
        ).execute()
        updated_cells = int(update_result.get("totalUpdatedCells") or 0)

    return {
        "spreadsheet_id": spreadsheet_id,
        "spreadsheet_title": spreadsheet.get("properties", {}).get("title"),
        "target_year": target_year,
        "target_month": target_month,
        "prepared_cells": len(update_data),
        "updated_cells": updated_cells,
        "dry_run": dry_run,
        "skipped_sheets": skipped_sheets,
        "post_content": post_content_summary,
    }


def sync_marketing_report_google_sheet() -> Any:
    try:
        target_year, target_month, target_start, target_end = _report_previous_completed_month()
        refresh_arg = (request.args.get("refresh") or "true").strip().lower()
        dry_run = (request.args.get("dry_run") or "false").strip().lower() in {"1", "true", "yes", "on"}
        refresh_result = None
        if refresh_arg not in {"0", "false", "no", "off"}:
            refresh_result = _refresh_report_month_insights(target_start, target_end)
        metadata = update_marketing_report_google_sheet(
            target_year=target_year,
            target_month=target_month,
            target_start=target_start,
            target_end=target_end,
            dry_run=dry_run,
        )
        return jsonify({
            "message": "Google marketing report sync prepared." if dry_run else "Google marketing report updated.",
            "refresh": refresh_result,
            **metadata,
        })
    except RuntimeError as error:
        return jsonify({"error": str(error)}), 500


def build_marketing_report_workbook(
    target_year: int | None = None,
    target_month: int | None = None,
    target_start: datetime | None = None,
    target_end: datetime | None = None,
) -> tuple[BytesIO, str, dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl is required to export the Excel marketing report.") from error

    template_path = _report_template_path()
    if not target_year or not target_month or not target_start or not target_end:
        target_year, target_month, target_start, target_end = _report_previous_completed_month()
    workbook = load_workbook(template_path)
    pages_by_key = _report_pages_by_key()
    account_metric_cache: dict[int, dict[str, int]] = {}
    filled_cells = 0
    skipped_sheets: list[str] = []

    for worksheet in workbook.worksheets:
        page = _report_page_for_sheet(worksheet.title, pages_by_key)
        if not page:
            skipped_sheets.append(worksheet.title)
            continue
        month_columns = _report_month_columns(worksheet)
        social_row = _report_social_row(worksheet)
        if not month_columns or not social_row:
            skipped_sheets.append(worksheet.title)
            continue
        target_column = month_columns.get(target_month)
        if not target_column:
            skipped_sheets.append(worksheet.title)
            continue

        platform_rows = _report_platform_rows(worksheet, social_row)
        for platform, platform_row in platform_rows.items():
            account = _report_account_for_platform(page, platform)
            if not account:
                continue
            if account.id not in account_metric_cache:
                account_metric_cache[account.id] = _report_account_month_metric_values(account, target_start, target_end)
            metric_values = account_metric_cache[account.id]
            for metric, row_number in _report_metric_rows(worksheet, platform_row).items():
                if metric not in metric_values:
                    continue
                worksheet.cell(row_number, target_column).value = metric_values[metric]
                filled_cells += 1

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"{template_path.stem} - exported.xlsx"
    return output, filename, {
        "filled_cells": filled_cells,
        "skipped_sheets": skipped_sheets,
        "target_year": target_year,
        "target_month": target_month,
    }


def export_marketing_report_workbook() -> Any:
    try:
        target_year, target_month, target_start, target_end = _report_previous_completed_month()
        refresh_arg = (request.args.get("refresh") or "true").strip().lower()
        if refresh_arg not in {"0", "false", "no", "off"}:
            _refresh_report_month_insights(target_start, target_end)
        output, filename, _metadata = build_marketing_report_workbook(
            target_year=target_year,
            target_month=target_month,
            target_start=target_start,
            target_end=target_end,
        )
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except RuntimeError as error:
        return jsonify({"error": str(error)}), 500


def _post_engagement_total(metrics: dict[str, Any]) -> int:
    for preferred in ("engagement", "total_interactions"):
        if metrics.get(preferred) is not None:
            return int(round(float(metrics.get(preferred) or 0)))
    return int(
        round(
            sum(
                float(metrics.get(name) or 0)
                for name in ["likes", "reactions", "comments", "shares", "saved"]
            )
        )
    )


def _post_row_from_reference(reference: PlatformPostReference) -> dict[str, Any]:
    metrics = {
        snapshot.metric_name: snapshot.metric_value
        for snapshot in PostInsightSnapshot.query.filter_by(
            internal_post_id=reference.internal_post_id,
            social_account_id=reference.social_account_id,
            platform_post_id=reference.platform_post_id,
        ).all()
        if snapshot.metric_value is not None and snapshot.status in {"ok", "stale"}
    }
    post = reference.post
    account = reference.account
    page = account.page if account and account.page else post.page if post and post.page else None
    thumbnail = post.media_list()[0] if post and post.media_list() else None
    return {
        "id": reference.id,
        "internal_post_id": reference.internal_post_id,
        "social_account_id": reference.social_account_id,
        "page_id": page.id if page else account.page_id if account else post.page_id if post else None,
        "page_name": page.name if page else None,
        "account_name": account.account_name if account else None,
        "thumbnail": build_local_media_url(thumbnail) if thumbnail else None,
        "caption": reference.caption_preview or (post.content[:280] if post and post.content else ""),
        "platform": reference.platform,
        "platform_post_id": reference.platform_post_id,
        "published_at": reference.published_at.isoformat() if reference.published_at else None,
        "views": int(metrics.get("views") or 0),
        "reach": int(metrics.get("reach") or 0),
        "engagement": _post_engagement_total(metrics),
        "comments": int(metrics.get("comments") or 0),
        "shares": int(metrics.get("shares") or 0),
        "permalink": reference.permalink,
        "state": "ready" if metrics else "No post insights yet",
        "metrics": metrics,
        "source": "saved_reference",
    }


def _recent_instagram_media_items(account: SocialAccount, *, since: datetime, limit: int) -> list[dict[str, Any]]:
    if account.platform != "instagram" or not account.page_id_external or not account.access_token:
        return []

    items: list[dict[str, Any]] = []
    after: str | None = None
    until = utcnow()
    while len(items) < limit:
        params: dict[str, Any] = {
            "fields": INSTAGRAM_MEDIA_FIELDS,
            "since": core.local_datetime_to_unix_timestamp(since),
            "until": core.local_datetime_to_unix_timestamp(until),
            "limit": min(max(limit - len(items), 1), 100),
        }
        if after:
            params["after"] = after
        payload = _graph_get(f"{account.page_id_external}/media", account.access_token, params)
        data = payload.get("data") if isinstance(payload, dict) else []
        if not isinstance(data, list) or not data:
            break
        for item in data:
            if isinstance(item, dict):
                items.append(item)
                if len(items) >= limit:
                    break
        paging = payload.get("paging") if isinstance(payload, dict) else None
        cursors = paging.get("cursors") if isinstance(paging, dict) else None
        next_after = cursors.get("after") if isinstance(cursors, dict) else None
        if not next_after or next_after == after:
            break
        after = str(next_after)
    return items


def _recent_instagram_media_rows(
    account: SocialAccount,
    *,
    since: datetime,
    limit: int,
    skipped_media_ids: set[str],
) -> list[PostInsightSnapshot]:
    stored: list[PostInsightSnapshot] = []
    try:
        media_items = _recent_instagram_media_items(account, since=since, limit=limit)
    except Exception as error:
        logger.info(
            "Instagram media discovery unavailable account_id=%s error=%s",
            account.id,
            _safe_error(error),
        )
        return stored

    for media in media_items:
        media_id = str(media.get("id") or "").strip()
        if not media_id or media_id in skipped_media_ids:
            continue
        published_at = _parse_graph_datetime(media.get("timestamp"))
        if published_at and published_at < since:
            continue
        metrics = _instagram_media_field_metrics(media)
        metrics.update(_fetch_instagram_media_metrics(account, media_id))
        reference = _upsert_remote_platform_reference(
            account,
            platform="instagram",
            platform_post_id=media_id,
            permalink=media.get("permalink"),
            published_at=published_at,
            media_type=media.get("media_type"),
            caption=str(media.get("caption") or "")[:280],
            thumbnail=media.get("thumbnail_url") or media.get("media_url"),
        )
        if reference:
            stored.extend(_store_post_metric_values(reference, metrics))
    return stored


def _recent_facebook_post_items(account: SocialAccount, *, since: datetime, limit: int) -> list[dict[str, Any]]:
    if account.platform != "facebook" or not account.page_id_external or not account.access_token:
        return []

    items: list[dict[str, Any]] = []
    after: str | None = None
    until = utcnow()
    while len(items) < limit:
        params: dict[str, Any] = {
            "fields": FACEBOOK_POST_FIELDS,
            "since": core.local_datetime_to_unix_timestamp(since),
            "until": core.local_datetime_to_unix_timestamp(until),
            "limit": min(max(limit - len(items), 1), 100),
        }
        if after:
            params["after"] = after
        payload = _graph_get(f"{account.page_id_external}/posts", account.access_token, params)
        data = payload.get("data") if isinstance(payload, dict) else []
        if not isinstance(data, list) or not data:
            break
        for item in data:
            if isinstance(item, dict):
                items.append(item)
                if len(items) >= limit:
                    break
        paging = payload.get("paging") if isinstance(payload, dict) else None
        cursors = paging.get("cursors") if isinstance(paging, dict) else None
        next_after = cursors.get("after") if isinstance(cursors, dict) else None
        if not next_after or next_after == after:
            break
        after = str(next_after)
    return items


def _recent_facebook_post_rows(
    account: SocialAccount,
    *,
    since: datetime,
    limit: int,
    skipped_post_ids: set[str],
) -> list[PostInsightSnapshot]:
    stored: list[PostInsightSnapshot] = []
    try:
        post_items = _recent_facebook_post_items(account, since=since, limit=limit)
    except Exception as error:
        logger.info(
            "Facebook post discovery unavailable account_id=%s error=%s",
            account.id,
            _safe_error(error),
        )
        return stored

    for item in post_items:
        post_id = str(item.get("id") or "").strip()
        if not post_id or post_id in skipped_post_ids:
            continue
        published_at = _parse_graph_datetime(item.get("created_time"))
        if published_at and published_at < since:
            continue
        caption = str(item.get("message") or item.get("story") or "")[:280]
        permalink = item.get("permalink_url")
        metrics = _facebook_post_field_metrics(item)
        metrics.update(_fetch_facebook_post_insight_metrics(account, post_id))
        if "engagement" not in metrics and any(key in metrics for key in ("comments", "shares", "reactions", "clicks")):
            metrics["engagement"] = sum(float(metrics.get(key) or 0) for key in ("comments", "shares", "reactions", "clicks"))
        reference = _upsert_remote_platform_reference(
            account,
            platform="facebook",
            platform_post_id=post_id,
            permalink=permalink,
            published_at=published_at,
            media_type=item.get("type"),
            caption=caption,
            thumbnail=_facebook_post_thumbnail(item),
        )
        if reference:
            stored.extend(_store_post_metric_values(reference, metrics))
    return stored


def _top_post_rows(limit: int = 10, *, days: int = POST_INSIGHTS_LOOKBACK_DAYS) -> list[dict[str, Any]]:
    limit = min(max(int(limit or 25), 1), 2000)
    accounts = _filtered_accounts()
    for account in accounts:
        _ensure_post_references_for_account(account)
    db.session.flush()

    account_ids = [account.id for account in accounts]
    if not account_ids:
        return []

    since = utcnow() - timedelta(days=max(days, 1))
    references = (
        PlatformPostReference.query.filter(PlatformPostReference.social_account_id.in_(account_ids))
        .order_by(PlatformPostReference.published_at.desc().nullslast(), PlatformPostReference.id.desc())
        .limit(max(limit * 2, 100))
        .all()
    )
    rows: list[dict[str, Any]] = []
    for reference in references:
        if reference.published_at and reference.published_at < since:
            continue
        rows.append(_post_row_from_reference(reference))

    deduped: list[dict[str, Any]] = []
    seen_remote_ids: set[str] = set()
    for row in rows:
        platform_post_id = str(row.get("platform_post_id") or "").strip()
        remote_key = f"{row.get('platform')}-{platform_post_id}"
        if platform_post_id and remote_key in seen_remote_ids:
            continue
        if platform_post_id:
            seen_remote_ids.add(remote_key)
        deduped.append(row)

    def row_date(row: dict[str, Any]) -> datetime:
        parsed = _parse_graph_datetime(row.get("published_at"))
        return parsed or datetime.max

    deduped.sort(key=lambda row: (row_date(row), str(row.get("platform") or ""), str(row.get("id") or "")))
    return deduped[:limit]


def get_analytics_summary() -> Any:
    accounts = _filtered_accounts()
    account_rows: list[dict[str, Any]] = []
    totals = {"views": 0, "engagement": 0, "followers": 0, "reach": 0, "visits": 0, "media_count": 0}
    for account in accounts:
        insights = _filtered_visible_insights(account)
        metrics = _metric_summary_from_insights(insights)
        for key in totals:
            totals[key] += metrics.get(key, 0)
        account_rows.append(
            {
                "id": account.id,
                "name": account.account_name or account.page.name if account.page else account.account_name or f"{account.platform} account",
                "platform": account.platform,
                "page_name": account.page.name if account.page else None,
                "followers": metrics.get("followers", 0),
                "views": metrics.get("views", 0),
                "engagement": metrics.get("engagement", 0),
                "reach": metrics.get("reach", 0),
                "visits": metrics.get("visits", 0),
                "last_refreshed": max((insight.refreshed_at for insight in insights), default=None).isoformat() if insights else None,
                "state": "warning" if account.test_error else "ready" if account.is_active and account.page_id_external and account.access_token else "needs_setup",
            }
        )

    best_account = max(account_rows, key=lambda row: row["views"] + row["engagement"], default=None)
    top_posts = _top_post_rows(limit=6)
    summary_cards = [
        {"label": "Total views", "value": totals["views"], "delta": None, "tone": "info"},
        {"label": "Content interactions", "value": totals["engagement"], "delta": None, "tone": "good"},
        {"label": "Total followers", "value": totals["followers"], "delta": None, "tone": "info"},
        {"label": "Follower growth", "value": 0, "delta": "tracking from next refresh", "tone": "neutral"},
        {"label": "Best account", "value": best_account["name"] if best_account else "No data", "delta": None, "tone": "good" if best_account else "neutral"},
        {"label": "Accounts needing attention", "value": sum(1 for row in account_rows if row["state"] != "ready"), "delta": None, "tone": "warn"},
    ]
    return jsonify(
        {
            "summary_cards": summary_cards,
            "totals": totals,
            "trends": _trend_rows(accounts),
            "account_comparison": account_rows,
            "top_posts": top_posts,
            "refresh": _scheduler_refresh_info(),
        }
    )


def get_analytics_trends(account_id: int | None = None) -> Any:
    accounts = [SocialAccount.query.get_or_404(account_id)] if account_id else _filtered_accounts()
    return jsonify({"items": _trend_rows(accounts)})


def get_analytics_posts() -> Any:
    limit = request.args.get("limit", 500, type=int) or 500
    days = request.args.get("days", POST_INSIGHTS_LOOKBACK_DAYS, type=int) or POST_INSIGHTS_LOOKBACK_DAYS
    return jsonify({"items": _top_post_rows(limit=limit, days=days)})


def get_analytics_post(post_id: int) -> Any:
    post = Post.query.get_or_404(post_id)
    references = PlatformPostReference.query.filter_by(internal_post_id=post.id).all()
    return jsonify(
        {
            "id": post.id,
            "page_id": post.page_id,
            "page_name": post.page.name if post.page else None,
            "caption": post.content or "",
            "media_type": post.media_type,
            "published_at": post.posted_at.isoformat() if post.posted_at else None,
            "status": post.status,
            "references": [
                {
                    **reference.to_dict(),
                    "insights": [
                        snapshot.to_dict()
                        for snapshot in PostInsightSnapshot.query.filter_by(
                            internal_post_id=post.id,
                            social_account_id=reference.social_account_id,
                            platform_post_id=reference.platform_post_id,
                        )
                        .order_by(PostInsightSnapshot.metric_name.asc())
                        .all()
                    ],
                }
                for reference in references
            ],
        }
    )


def get_analytics_raw() -> Any:
    limit = min(max(request.args.get("limit", 250, type=int) or 250, 1), 1000)
    account_ids = [account.id for account in _filtered_accounts()]
    rows: list[dict[str, Any]] = []
    if account_ids:
        snapshots = (
            AccountInsightSnapshot.query.filter(AccountInsightSnapshot.social_account_id.in_(account_ids))
            .order_by(AccountInsightSnapshot.fetched_at.desc())
            .limit(limit)
            .all()
        )
        rows.extend(
            {
                **snapshot.to_dict(),
                "record_type": "account",
                "friendly_label": _friendly_metric_label(snapshot.metric_name),
            }
            for snapshot in snapshots
        )
        remaining = max(limit - len(rows), 0)
        if remaining:
            post_snapshots = (
                PostInsightSnapshot.query.filter(PostInsightSnapshot.social_account_id.in_(account_ids))
                .order_by(PostInsightSnapshot.fetched_at.desc())
                .limit(remaining)
                .all()
            )
            rows.extend(
                {
                    **snapshot.to_dict(),
                    "record_type": "post",
                    "friendly_label": _friendly_metric_label(snapshot.metric_name),
                }
                for snapshot in post_snapshots
            )
    return jsonify({"items": rows})


def get_analytics_accounts() -> Any:
    platform = (request.args.get("platform") or "all").strip().lower()
    query = SocialAccount.query.filter(SocialAccount.platform.in_(["facebook", "instagram"]))
    if platform in {"facebook", "instagram"}:
        query = query.filter(SocialAccount.platform == platform)
    accounts = query.order_by(SocialAccount.platform.asc(), SocialAccount.account_name.asc()).all()
    return jsonify({"items": [_account_payload(account) for account in accounts]})


def get_analytics_account(account_id: int) -> Any:
    account = SocialAccount.query.get_or_404(account_id)
    return jsonify(_account_payload(account))


def get_account_insights(account_id: int) -> Any:
    account = SocialAccount.query.get_or_404(account_id)
    insights = _latest_visible_insights_for_account(account.id)
    return jsonify({"items": [insight.to_dict() for insight in insights]})


def _analytics_refresh_state_payload() -> dict[str, Any]:
    with _analytics_refresh_lock:
        payload = dict(_analytics_refresh_state)
    for key in ("started_at", "finished_at"):
        value = payload.get(key)
        if isinstance(value, datetime):
            payload[key] = value.isoformat()
    return payload


def _set_analytics_refresh_state(**updates: Any) -> None:
    with _analytics_refresh_lock:
        _analytics_refresh_state.update(updates)


def _run_analytics_refresh_job(
    job_id: str,
    account_id: int | None,
    force: bool,
    start: datetime,
    end: datetime,
) -> None:
    with core.app.app_context():
        try:
            if account_id:
                account = SocialAccount.query.get(account_id)
                if not account:
                    raise RuntimeError(f"Analytics account {account_id} was not found.")
                _set_analytics_refresh_state(
                    status="running",
                    message=f"Refreshing {account.platform} account {account.account_name or account.id}.",
                    progress_current=1,
                    progress_total=1,
                )
                logger.info("Manual analytics refresh job %s started for account_id=%s.", job_id, account_id)
                result = refresh_account_insights(account, force=force, start=start, end=end)
            else:
                def update_progress(current: int, total: int, account: SocialAccount) -> None:
                    _set_analytics_refresh_state(
                        status="running",
                        message=f"Refreshing all accounts: {account.platform} account {account.account_name or account.id} ({current}/{total}).",
                        progress_current=current,
                        progress_total=total,
                    )

                logger.info("Manual analytics refresh job %s started for all accounts.", job_id)
                result = refresh_all_social_insights(force=force, paced=True, start=start, end=end, progress_callback=update_progress)

            _set_analytics_refresh_state(
                status="finished",
                message="Analytics refresh finished.",
                finished_at=utcnow(),
                result=result,
                error=None,
            )
            logger.info("Manual analytics refresh job %s finished: %s", job_id, result)
        except Exception as error:
            db.session.rollback()
            message = _safe_error(error)
            _set_analytics_refresh_state(
                status="failed",
                message="Analytics refresh failed.",
                finished_at=utcnow(),
                error=message,
            )
            logger.exception("Manual analytics refresh job %s failed: %s", job_id, message)


def get_analytics_refresh_status() -> Any:
    return jsonify(_analytics_refresh_state_payload())


def refresh_analytics() -> Any:
    account_id = request.args.get("account_id", type=int)
    force = str(request.args.get("force") or "").lower() in {"1", "true", "yes", "on"}
    try:
        start, end = _parse_refresh_range_args()
    except Exception as error:
        return jsonify({"status": "failed", "accepted": False, "message": _safe_error(error), "error": _safe_error(error)}), 400
    progress_total = 1 if account_id else SocialAccount.query.filter(
        SocialAccount.platform.in_(["facebook", "instagram"]),
        SocialAccount.is_active.is_(True),
    ).count()
    with _analytics_refresh_lock:
        if _analytics_refresh_state.get("status") in {"queued", "running"}:
            payload = dict(_analytics_refresh_state)
            payload["accepted"] = False
            payload["message"] = payload.get("message") or "Analytics refresh is already running."
            for key in ("started_at", "finished_at"):
                value = payload.get(key)
                if isinstance(value, datetime):
                    payload[key] = value.isoformat()
            return jsonify(payload), 202

        job_id = uuid.uuid4().hex
        _analytics_refresh_state.update(
            {
                "id": job_id,
                "status": "queued",
                "message": "Analytics refresh queued for one account." if account_id else "Analytics refresh queued for all connected Facebook and Instagram accounts.",
                "account_id": account_id,
                "started_at": utcnow(),
                "finished_at": None,
                "progress_current": 0,
                "progress_total": progress_total,
                "range_start": start.isoformat(),
                "range_end": end.isoformat(),
                "result": None,
                "error": None,
            }
        )

    thread = Thread(target=_run_analytics_refresh_job, args=(job_id, account_id, force, start, end), daemon=True)
    thread.start()
    payload = _analytics_refresh_state_payload()
    payload["accepted"] = True
    return jsonify(payload), 202
